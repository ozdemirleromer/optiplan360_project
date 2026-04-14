import sys
import tempfile
from pathlib import Path
import shutil
import base64

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.database import Base
from app.constants.optiplan_workflow import ExportContractRules
from app.exceptions import BusinessRuleError, ValidationError
from app.models.optiplan_workflow import OptiPlanWorkflowSatir
from app.services.gemini_ocr_adapter import OcrDocumentResult
from app.services.optiplan_workflow_service import EXPORT_COLUMNS, optiplan_workflow_service
import app.models  # noqa: F401




def _case_dir(case_path: Path, label: str) -> Path:
    return case_path.parent / f"{case_path.name}_{label}"


def _valid_png_bytes() -> bytes:
    return base64.b64decode("iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+X2XQAAAAASUVORK5CYII=")


def _configure_settings(db_session, tmp_path: Path, *, watcher_aktif_mi: bool = True) -> dict:
    paths = {
        "whatsapp_raw_klasoru": str(_case_dir(tmp_path, "whatsapp_raw")),
        "scanner_raw_klasoru": str(_case_dir(tmp_path, "scanner_raw")),
        "manuel_raw_klasoru": str(_case_dir(tmp_path, "manuel_raw")),
        "email_raw_klasoru": str(_case_dir(tmp_path, "email_raw")),
        "islenmis_klasoru": str(_case_dir(tmp_path, "islenmis")),
        "arsiv_klasoru": str(_case_dir(tmp_path, "arsiv")),
        "xml_okuma_klasoru": str(_case_dir(tmp_path, "xml")),
        "xlsx_cikti_klasoru": str(_case_dir(tmp_path, "xlsx")),
        "hatali_klasoru": str(_case_dir(tmp_path, "hatali")),
        "fis_evrak_no_formati": "SIP-{seq:06d}",
        "arsiv_zaman_damgasi_formati": "%Y%m%d_%H%M%S",
        "xlsx_aktif_mi": True,
        "watcher_aktif_mi": watcher_aktif_mi,
        "yeniden_deneme_sayisi": 2,
    }
    return optiplan_workflow_service.update_folder_settings(db_session, paths)


def _create_record_with_row(db_session, tmp_path: Path) -> tuple[str, str]:
    _configure_settings(db_session, tmp_path)
    record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="ornek.jpg",
        content=b"ornek-gorsel",
        kaynak_klasor="manuel_raw",
    )
    row = OptiPlanWorkflowSatir(
        id="row-1",
        kayit_uuid=record["kayit_uuid"],
        satir_sirasi=1,
        boy=100,
        en=50,
        adet=1,
        grain=3,
        satir_kaynagi="OCR",
    )
    db_session.add(row)
    db_session.commit()
    return record["kayit_uuid"], row.id


def test_folder_settings_singleton_round_trip(db_session, workspace_tmp_path: Path):
    saved = _configure_settings(db_session, workspace_tmp_path)
    fetched = optiplan_workflow_service.serialize_folder_settings(
        optiplan_workflow_service.get_folder_settings(db_session)
    )

    assert fetched["manuel_raw_klasoru"] == saved["manuel_raw_klasoru"]
    assert fetched["watcher_aktif_mi"] is True


def test_manual_import_moves_to_archive_and_blocks_duplicate(db_session, workspace_tmp_path: Path):
    settings = _configure_settings(db_session, workspace_tmp_path)

    first_record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="ornek.png",
        content=b"ayni-icerik",
        kaynak_klasor="manuel_raw",
    )

    archive_files = list(Path(settings["arsiv_klasoru"]).iterdir())
    assert first_record["kaynak_klasor"] == "manuel_raw"
    assert len(archive_files) == 1
    assert not list(Path(settings["manuel_raw_klasoru"]).iterdir())
    assert Path(settings["islenmis_klasoru"]).exists()

    with pytest.raises(BusinessRuleError):
        optiplan_workflow_service.manual_import(
            db_session,
            file_name="tekrar.png",
            content=b"ayni-icerik",
            kaynak_klasor="manuel_raw",
        )

    second_record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="tekrar.png",
        content=b"ayni-icerik",
        kaynak_klasor="manuel_raw",
        force_duplicate=True,
    )
    assert second_record["kayit_uuid"] != first_record["kayit_uuid"]
    assert len(list(Path(settings["arsiv_klasoru"]).iterdir())) == 2


def test_manual_import_generates_ocr_rows_for_valid_image(db_session, workspace_tmp_path: Path):
    _configure_settings(db_session, workspace_tmp_path)

    record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="ocr-kanit.png",
        content=_valid_png_bytes(),
        kaynak_klasor="manuel_raw",
    )

    assert record["dosya_durumu"] == "PHASE_2_OCR_KONTROL"
    assert record["aktif_faz"] == 2
    assert record["okunan_cari_unvan"] == "ABC Mobilya"
    assert record["okunan_cari_telefon"] == "5321234567"
    assert record["malzeme"] == "Beyaz MDFLAM"
    assert len(record["satirlar"]) == 3
    assert record["satirlar"][0]["boy"] == 700
    assert record["satirlar"][0]["en"] == 400
    assert record["satirlar"][0]["adet"] == 2
    assert record["satirlar"][0]["satir_kaynagi"] == "OCR"
    assert record["satirlar"][0]["hucre_guven_skorlari"]["boy"] == 85.0
    assert any(item["alan_adi"] == "ocr_ingest_completed" for item in record["audit_kayitlari"])



def test_manual_import_falls_back_when_gemini_returns_no_rows(db_session, workspace_tmp_path: Path, monkeypatch):
    _configure_settings(db_session, workspace_tmp_path)

    def _empty_gemini_result(*args, **kwargs):
        return OcrDocumentResult(
            model_name="gemini-test",
            latency_ms=123,
            raw_response_text="",
            parse_error="empty rows",
        )

    monkeypatch.setattr("app.services.optiplan_workflow_service.run_gemini_ocr", _empty_gemini_result)

    record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="ocr-gemini-empty.png",
        content=_valid_png_bytes(),
        kaynak_klasor="manuel_raw",
    )

    assert record["dosya_durumu"] == "PHASE_2_OCR_KONTROL"
    assert len(record["satirlar"]) == 3
    assert record["ocr_ham_json"]["engine"] == "simulated_local_ocr"
    assert record["ocr_ham_json"]["status"] == "COMPLETED"


def test_scan_watch_folders_returns_empty_when_watcher_disabled(db_session, workspace_tmp_path: Path):
    settings = _configure_settings(db_session, workspace_tmp_path, watcher_aktif_mi=False)
    source_file = Path(settings["whatsapp_raw_klasoru"]) / "bekleyen.jpg"
    source_file.parent.mkdir(parents=True, exist_ok=True)
    source_file.write_bytes(b"bekleyen")

    assert optiplan_workflow_service.scan_watch_folders(db_session) == []
    assert source_file.exists()


def test_scan_watch_folders_generates_ocr_rows_for_valid_image(db_session, workspace_tmp_path: Path):
    settings = _configure_settings(db_session, workspace_tmp_path, watcher_aktif_mi=True)
    source_file = Path(settings["scanner_raw_klasoru"]) / "scanner-kanit.png"
    source_file.parent.mkdir(parents=True, exist_ok=True)
    source_file.write_bytes(_valid_png_bytes())

    records = optiplan_workflow_service.scan_watch_folders(db_session)

    assert len(records) == 1
    assert records[0]["kaynak_klasor"] == "scanner_raw"
    assert records[0]["dosya_durumu"] == "PHASE_2_OCR_KONTROL"
    assert len(records[0]["satirlar"]) == 3
    assert not source_file.exists()


def test_phase2_update_creates_audit_and_remove_restore_round_trip(db_session, workspace_tmp_path: Path):
    kayit_uuid, row_id = _create_record_with_row(db_session, workspace_tmp_path)

    updated = optiplan_workflow_service.update_phase2(
        db_session,
        kayit_uuid,
        rows=[
            {
                "id": row_id,
                "boy": 110,
                "en": 60,
                "adet": 2,
                "hucre_guven_skorlari": {"boy": 75},
                "satir_guven_skor_ozeti": {"min": 75},
            }
        ],
        metadata={"okunan_cari_unvan": "Ornek Cari"},
    )

    assert updated["aktif_faz"] == 2
    assert len(updated["audit_kayitlari"]) == 3
    assert updated["satirlar"][0]["boy"] == 110

    removed = optiplan_workflow_service.remove_phase2_row(db_session, kayit_uuid, row_id)
    assert removed["satirlar"] == []
    assert len(removed["cikarilan_satirlar"]) == 1
    removed_audit = [row for row in removed["audit_kayitlari"] if row["alan_adi"] == "phase2_row_removed"]
    assert len(removed_audit) == 1
    assert row_id in (removed_audit[0]["eski_deger"] or "")

    restored = optiplan_workflow_service.restore_phase2_row(
        db_session,
        kayit_uuid,
        removed["cikarilan_satirlar"][0]["id"],
    )
    assert len(restored["satirlar"]) == 1
    assert restored["cikarilan_satirlar"] == []
    assert restored["satirlar"][0]["id"] == row_id
    restored_audit = [row for row in restored["audit_kayitlari"] if row["alan_adi"] == "phase2_row_restored"]
    assert len(restored_audit) == 1
    assert row_id in (restored_audit[0]["yeni_deger"] or "")


def test_phase2_update_accepts_camel_case_score_aliases(db_session, workspace_tmp_path: Path):
    kayit_uuid, row_id = _create_record_with_row(db_session, workspace_tmp_path)

    updated = optiplan_workflow_service.update_phase2(
        db_session,
        kayit_uuid,
        rows=[
            {
                "id": row_id,
                "boy": 115,
                "en": 65,
                "adet": 3,
                "malzeme": "18MM Beyaz",
                "grain": 3,
                "bilgi": "Arka yuz",
                "delik1": "12",
                "hucreGuvenSkorlari": {
                    "boy": 85,
                    "en": 90,
                    "adet": 95,
                    "malzeme": 92,
                    "grain": 91,
                    "bilgi": 89,
                    "delik1": 81,
                },
                "satirGuvenSkorOzeti": {"onaylanan_hucreler": ["boy", "delik_1"]},
            }
        ],
        metadata={"okunan_cari_unvan": "Alias Cari"},
    )

    row = updated["satirlar"][0]
    assert row["boy"] == 115
    assert row["en"] == 65
    assert row["adet"] == 3
    assert row["malzeme"] == "18MM Beyaz"
    assert row["grain"] == 3
    assert row["bilgi"] == "Arka yuz"
    assert row["delik_1"] == "12"
    assert row["hucre_guven_skorlari"]["boy"] == 85
    assert row["hucre_guven_skorlari"]["delik1"] == 81
    assert row["satir_guven_skor_ozeti"]["onaylanan_hucreler"] == ["boy", "delik_1"]


def test_approve_phase2_rejects_unapproved_low_confidence_non_core_field(db_session, workspace_tmp_path: Path):
    kayit_uuid, row_id = _create_record_with_row(db_session, workspace_tmp_path)

    optiplan_workflow_service.update_phase2(
        db_session,
        kayit_uuid,
        rows=[
            {
                "id": row_id,
                "boy": 115,
                "en": 65,
                "adet": 3,
                "malzeme": "18MM Beyaz",
                "grain": 3,
                "bilgi": "Arka yuz",
                "delik1": "12",
                "hucreGuvenSkorlari": {
                    "boy": 95,
                    "en": 94,
                    "adet": 96,
                    "malzeme": 55,
                    "grain": 91,
                    "bilgi": 92,
                    "delik1": 90,
                },
            }
        ],
        metadata={"okunan_cari_unvan": "Gate Cari"},
    )

    with pytest.raises(ValidationError, match="alan=malzeme"):
        optiplan_workflow_service.approve_phase2(db_session, kayit_uuid)


def test_approve_phase2_allows_transition_after_explicit_operator_approval(db_session, workspace_tmp_path: Path):
    kayit_uuid, row_id = _create_record_with_row(db_session, workspace_tmp_path)

    optiplan_workflow_service.update_phase2(
        db_session,
        kayit_uuid,
        rows=[
            {
                "id": row_id,
                "boy": 115,
                "en": 65,
                "adet": 3,
                "malzeme": "18MM Beyaz",
                "grain": 3,
                "bilgi": "Arka yuz",
                "delik1": "12",
                "hucreGuvenSkorlari": {
                    "boy": 72,
                    "en": 94,
                    "adet": 96,
                    "malzeme": 55,
                    "grain": 91,
                    "bilgi": 92,
                    "delik1": 78,
                },
                "satirGuvenSkorOzeti": {"onaylanan_hucreler": ["malzeme", "delik_1"]},
            }
        ],
        metadata={"okunan_cari_unvan": "Approved Cari"},
    )

    row = db_session.query(OptiPlanWorkflowSatir).filter(OptiPlanWorkflowSatir.id == row_id).one()
    row.boy_onay = "ONAYLANDI"
    db_session.commit()

    approved = optiplan_workflow_service.approve_phase2(db_session, kayit_uuid)

    assert approved["aktif_faz"] == 3
    assert approved["dosya_durumu"] == "PHASE_3_SIPARIS_DUZENLEME"
    phase_transition_audit = [row for row in approved["audit_kayitlari"] if row["alan_adi"] == "phase_transition"]
    assert len(phase_transition_audit) == 1


def test_phase3_update_creates_persistent_audit_entries_for_record_and_row_changes(
    db_session,
    workspace_tmp_path: Path,
):
    kayit_uuid, row_id = _create_record_with_row(db_session, workspace_tmp_path)

    updated = optiplan_workflow_service.update_phase3(
        db_session,
        kayit_uuid,
        {
            "cari_unvan": "Deneme Cari",
            "cari_kodu": "CR-001",
            "siparis_no": "SIP-001",
            "termin": "2026-03-14",
            "teslim_tarihi": "2026-03-15",
            "teslimat_adresi": "Test Sevkiyat Deposu",
            "odeme_sekli": "HAVALE",
            "malzeme": "Beyaz Mdflam",
            "stok_kodu": "STK-001",
            "rows": [
                {
                    "id": row_id,
                    "satir_sirasi": 1,
                    "malzeme": "Beyaz Mdflam",
                    "boy": 550,
                    "en": 320,
                    "adet": 2,
                    "grain": 3,
                    "satir_kaynagi": "MANUEL",
                }
            ],
        },
    )

    audit_fields = {row["alan_adi"] for row in updated["audit_kayitlari"]}
    assert "cari_unvan" in audit_fields
    assert "cari_kodu" in audit_fields
    assert "siparis_no" in audit_fields
    assert "termin" in audit_fields
    assert "phase3_row.boy" in audit_fields
    assert "phase3_row.en" in audit_fields
    assert "phase3_row.adet" in audit_fields
    row_audit = [row for row in updated["audit_kayitlari"] if row["alan_adi"] == "phase3_row.boy"]
    assert len(row_audit) == 1
    assert row_audit[0]["satir_id"] == row_id


def test_phase3_export_preview_and_xlsx_contract(db_session, workspace_tmp_path: Path):
    _configure_settings(db_session, workspace_tmp_path)
    record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="siparis.pdf",
        content=b"siparis-verisi",
        kaynak_klasor="email_raw",
    )

    updated = optiplan_workflow_service.update_phase3(
        db_session,
        record["kayit_uuid"],
        {
            "cari_unvan": "Özdemir Mobilya",
            "cari_kodu": "CARI001",
            "siparis_no": "SIP-000001",
            "termin": "2026-03-11",
            "teslim_tarihi": "2026-03-12",
            "teslimat_adresi": "Istanbul Ana Depo",
            "odeme_sekli": "HAVALE",
            "malzeme": "Beyaz Mdflam",
            "stok_kodu": "STK001",
            "bant_kalinligi": "0.40 MM",
            "grain_varsayilan": 3,
            "rows": [
                {
                    "satir_sirasi": 1,
                    "malzeme": "Beyaz Mdflam",
                    "boy": 500,
                    "en": 300,
                    "adet": 2,
                    "grain": 3,
                    "bilgi": "Kapak",
                    "u1": True,
                    "u2": False,
                    "k1": False,
                    "k2": False,
                    "delik_1": "123",
                    "delik_2": "456",
                    "satir_kaynagi": "OCR",
                },
                {
                    "satir_sirasi": 2,
                    "malzeme": "Beyaz Mdflam",
                    "boy": 500,
                    "en": 300,
                    "adet": 1,
                    "grain": 3,
                    "bilgi": "Kapak",
                    "u1": True,
                    "u2": False,
                    "k1": False,
                    "k2": False,
                    "delik_1": "123",
                    "delik_2": "456",
                    "satir_kaynagi": "MANUEL",
                },
            ],
        },
    )

    assert updated["aktif_faz"] == 3
    preview = optiplan_workflow_service.export_preview(db_session, record["kayit_uuid"], xlsx_aktif_mi=True)

    assert list(preview["satirlar"][0].keys()) == EXPORT_COLUMNS
    assert preview["satirlar"][0]["[P_MINQ]"] == 3
    assert preview["satirlar"][0]["[P_EDGE_MAT_UP]"] == "04"
    assert "OZDEMIR_MOBILYA" in preview["dosya_adi"]
    assert "BEYAZ_MDFLAM" in preview["dosya_adi"]
    exported = optiplan_workflow_service.export_record(db_session, record["kayit_uuid"], xlsx_aktif_mi=True)
    workbook = load_workbook(exported["generated_files"][0])
    sheet = workbook.active

    assert exported["generated_files"][0].endswith(".xlsx")
    assert exported["durum"] == "BASARILI"
    assert [cell.value for cell in sheet[1]] == EXPORT_COLUMNS
    assert sheet.cell(row=2, column=4).value == 3
    assert sheet.cell(row=2, column=7).value == "04"


def test_export_preview_advances_revision_for_reused_export_stem(db_session, workspace_tmp_path: Path):
    _configure_settings(db_session, workspace_tmp_path)

    def phase3_payload() -> dict:
        return {
            "cari_unvan": "Özdemir Mobilya",
            "cari_kodu": "CARI001",
            "siparis_no": "SIP-REV-001",
            "termin": "2026-03-11",
            "teslim_tarihi": "2026-03-12",
            "teslimat_adresi": "Istanbul Ana Depo",
            "odeme_sekli": "HAVALE",
            "malzeme": "Beyaz Mdflam",
            "stok_kodu": "STK001",
            "rows": [
                {
                    "satir_sirasi": 1,
                    "malzeme": "Beyaz Mdflam",
                    "boy": 500,
                    "en": 300,
                    "adet": 2,
                    "grain": 3,
                    "satir_kaynagi": "MANUEL",
                }
            ],
        }

    first_record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="ilk-is.pdf",
        content=b"ilk-export-verisi",
        kaynak_klasor="email_raw",
    )
    optiplan_workflow_service.update_phase3(
        db_session,
        first_record["kayit_uuid"],
        phase3_payload(),
    )
    first_export = optiplan_workflow_service.export_record(
        db_session,
        first_record["kayit_uuid"],
        xlsx_aktif_mi=True,
    )

    second_record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="ikinci-is.pdf",
        content=b"ikinci-export-verisi",
        kaynak_klasor="email_raw",
    )
    optiplan_workflow_service.update_phase3(
        db_session,
        second_record["kayit_uuid"],
        phase3_payload(),
    )
    second_preview = optiplan_workflow_service.export_preview(
        db_session,
        second_record["kayit_uuid"],
        xlsx_aktif_mi=True,
    )

    assert first_export["revizyon_no"] == 1
    assert second_preview["dosya_adi"] == f"{first_export['dosya_adi']}_v2"
    assert second_preview["revizyon_no"] == 2


def test_export_record_rejects_when_no_format_is_selected(db_session, workspace_tmp_path: Path):
    _configure_settings(db_session, workspace_tmp_path)
    record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="no-format.pdf",
        content=b"export-verisi",
        kaynak_klasor="email_raw",
    )

    optiplan_workflow_service.update_phase3(
        db_session,
        record["kayit_uuid"],
        {
            "cari_unvan": "Ornek Cari",
            "cari_kodu": "CARI001",
            "siparis_no": "SIP-NOFMT-001",
            "termin": "2026-03-12",
            "teslim_tarihi": "2026-03-13",
            "teslimat_adresi": "No Format Depo",
            "odeme_sekli": "HAVALE",
            "malzeme": "Beyaz Mdflam",
            "stok_kodu": "STK001",
            "rows": [
                {
                    "satir_sirasi": 1,
                    "malzeme": "Beyaz Mdflam",
                    "boy": 500,
                    "en": 300,
                    "adet": 1,
                    "grain": 3,
                    "satir_kaynagi": "MANUEL",
                }
            ],
        },
    )

    preview = optiplan_workflow_service.export_preview(
        db_session,
        record["kayit_uuid"],
        xlsx_aktif_mi=False,
    )

    assert preview["xlsx_aktif_mi"] is False
    assert preview["export_manifest"]["requested_formats"] == []

    with pytest.raises(ValidationError, match="En az bir export formati secilmelidir"):
        optiplan_workflow_service.export_record(
            db_session,
            record["kayit_uuid"],
            xlsx_aktif_mi=False,
        )


def test_export_record_writes_export_status_audit(db_session, workspace_tmp_path: Path):
    _configure_settings(db_session, workspace_tmp_path)
    record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="audit-export.pdf",
        content=b"audit-export",
        kaynak_klasor="email_raw",
    )

    optiplan_workflow_service.update_phase3(
        db_session,
        record["kayit_uuid"],
        {
            "cari_unvan": "Ornek Cari",
            "cari_kodu": "CARI001",
            "siparis_no": "SIP-AUDIT-001",
            "termin": "2026-03-12",
            "teslim_tarihi": "2026-03-13",
            "teslimat_adresi": "Audit Merkez",
            "odeme_sekli": "HAVALE",
            "malzeme": "Beyaz Mdflam",
            "stok_kodu": "STK001",
            "rows": [
                {
                    "satir_sirasi": 1,
                    "malzeme": "Beyaz Mdflam",
                    "boy": 500,
                    "en": 300,
                    "adet": 1,
                    "grain": 3,
                    "satir_kaynagi": "MANUEL",
                }
            ],
        },
    )

    exported = optiplan_workflow_service.export_record(
        db_session,
        record["kayit_uuid"],
        xlsx_aktif_mi=True,
    )
    detail = optiplan_workflow_service.get_record(db_session, record["kayit_uuid"])

    export_audit = [row for row in detail["audit_kayitlari"] if row["alan_adi"] == "export_durum"]

    assert exported["durum"] == "BASARILI"
    assert len(export_audit) == 1
    assert "durum=BASARILI" in (export_audit[0]["yeni_deger"] or "")
    assert "requested_formats=1" in (export_audit[0]["yeni_deger"] or "")
    assert "generated_formats=1" in (export_audit[0]["yeni_deger"] or "")

def test_export_record_fallbacks_and_logs_unknown_status_anomaly(
    db_session,
    workspace_tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _configure_settings(db_session, workspace_tmp_path)
    record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="anomaly-export.pdf",
        content=b"anomaly-export",
        kaynak_klasor="email_raw",
    )

    optiplan_workflow_service.update_phase3(
        db_session,
        record["kayit_uuid"],
        {
            "cari_unvan": "Ornek Cari",
            "cari_kodu": "CARI001",
            "siparis_no": "SIP-ANOM-001",
            "termin": "2026-03-12",
            "teslim_tarihi": "2026-03-13",
            "teslimat_adresi": "Anomali Sevkiyat",
            "odeme_sekli": "KREDI",
            "malzeme": "Beyaz Mdflam",
            "stok_kodu": "STK001",
            "rows": [
                {
                    "satir_sirasi": 1,
                    "malzeme": "Beyaz Mdflam",
                    "boy": 500,
                    "en": 300,
                    "adet": 1,
                    "grain": 3,
                    "satir_kaynagi": "MANUEL",
                }
            ],
        },
    )

    monkeypatch.setattr(
        ExportContractRules,
        "validate_export_status",
        staticmethod(lambda _status: False),
    )

    exported = optiplan_workflow_service.export_record(
        db_session,
        record["kayit_uuid"],
        xlsx_aktif_mi=True,
    )
    detail = optiplan_workflow_service.get_record(db_session, record["kayit_uuid"])

    anomaly_audit = [row for row in detail["audit_kayitlari"] if row["alan_adi"] == "export_durum_anomali"]

    assert exported["durum"] == "HATALI"
    assert len(anomaly_audit) == 1
    assert anomaly_audit[0]["eski_deger"] == "BASARILI"
    assert "fallback=HATALI" in (anomaly_audit[0]["yeni_deger"] or "")


def test_get_export_status_anomalies_returns_summary_and_items(
    db_session,
    workspace_tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _configure_settings(db_session, workspace_tmp_path)
    record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="telemetry-export.pdf",
        content=b"telemetry-export",
        kaynak_klasor="email_raw",
    )

    optiplan_workflow_service.update_phase3(
        db_session,
        record["kayit_uuid"],
        {
            "cari_unvan": "Ornek Cari",
            "cari_kodu": "CARI001",
            "siparis_no": "SIP-TELEMETRY-001",
            "termin": "2026-03-12",
            "teslim_tarihi": "2026-03-13",
            "teslimat_adresi": "Telemetry Depo",
            "odeme_sekli": "HAVALE",
            "malzeme": "Beyaz Mdflam",
            "stok_kodu": "STK001",
            "rows": [
                {
                    "satir_sirasi": 1,
                    "malzeme": "Beyaz Mdflam",
                    "boy": 500,
                    "en": 300,
                    "adet": 1,
                    "grain": 3,
                    "satir_kaynagi": "MANUEL",
                }
            ],
        },
    )

    monkeypatch.setattr(
        ExportContractRules,
        "validate_export_status",
        staticmethod(lambda _status: False),
    )

    optiplan_workflow_service.export_record(
        db_session,
        record["kayit_uuid"],
        xlsx_aktif_mi=True,
    )

    telemetry = optiplan_workflow_service.get_export_status_anomalies(db_session, limit=10)

    assert telemetry["limit"] == 10
    assert telemetry["summary"]["total_records"] >= 1
    assert telemetry["summary"]["distinct_records"] >= 1
    assert telemetry["summary"]["last_created_at"] is not None
    assert len(telemetry["items"]) >= 1
    assert telemetry["items"][0]["alan_adi"] == "export_durum_anomali"
    assert telemetry["summary"]["status_breakdown"].get("BASARILI", 0) >= 1


def test_get_export_status_anomalies_supports_filters_and_offset(
    db_session,
    workspace_tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    _configure_settings(db_session, workspace_tmp_path)
    record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="telemetry-filter-export.pdf",
        content=b"telemetry-filter-export",
        kaynak_klasor="email_raw",
    )

    optiplan_workflow_service.update_phase3(
        db_session,
        record["kayit_uuid"],
        {
            "cari_unvan": "Ornek Cari",
            "cari_kodu": "CARI001",
            "siparis_no": "SIP-TELEMETRY-002",
            "termin": "2026-03-12",
            "teslim_tarihi": "2026-03-13",
            "teslimat_adresi": "Telemetry Depo 2",
            "odeme_sekli": "HAVALE",
            "malzeme": "Beyaz Mdflam",
            "stok_kodu": "STK001",
            "rows": [
                {
                    "satir_sirasi": 1,
                    "malzeme": "Beyaz Mdflam",
                    "boy": 500,
                    "en": 300,
                    "adet": 1,
                    "grain": 3,
                    "satir_kaynagi": "MANUEL",
                }
            ],
        },
    )

    monkeypatch.setattr(
        ExportContractRules,
        "validate_export_status",
        staticmethod(lambda _status: False),
    )

    optiplan_workflow_service.export_record(
        db_session,
        record["kayit_uuid"],
        xlsx_aktif_mi=True,
    )

    telemetry = optiplan_workflow_service.get_export_status_anomalies(
        db_session,
        limit=1,
        offset=0,
        kayit_uuid=record["kayit_uuid"],
    )

    assert telemetry["limit"] == 1
    assert telemetry["offset"] == 0
    assert telemetry["filters"]["kayit_uuid"] == record["kayit_uuid"]
    assert len(telemetry["items"]) == 1
    assert telemetry["items"][0]["kayit_uuid"] == record["kayit_uuid"]


def test_get_export_status_anomalies_rejects_invalid_time_range(db_session):
    with pytest.raises(ValidationError):
        optiplan_workflow_service.get_export_status_anomalies(
            db_session,
            from_ts="2026-03-13T00:00:00+00:00",
            to_ts="2026-03-12T00:00:00+00:00",
        )


def test_get_export_status_anomalies_rejects_invalid_datetime_format(db_session):
    with pytest.raises(ValidationError):
        optiplan_workflow_service.get_export_status_anomalies(
            db_session,
            from_ts="2026/03/12 10:00:00",
        )


def test_get_export_status_anomalies_rejects_datetime_without_timezone(db_session):
    with pytest.raises(ValidationError):
        optiplan_workflow_service.get_export_status_anomalies(
            db_session,
            from_ts="2026-03-12T10:00:00",
        )


def test_export_preview_row_variants_cover_edges_and_holes(db_session, workspace_tmp_path: Path):
    _configure_settings(db_session, workspace_tmp_path)
    record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="varyantlar.pdf",
        content=b"varyant-verisi",
        kaynak_klasor="email_raw",
    )

    optiplan_workflow_service.update_phase3(
        db_session,
        record["kayit_uuid"],
        {
            "cari_unvan": "Ornek Cari",
            "cari_kodu": "CARI001",
            "siparis_no": "SIP-VAR-001",
            "termin": "2026-03-12",
            "teslim_tarihi": "2026-03-13",
            "teslimat_adresi": "Varyant Depo",
            "odeme_sekli": "HAVALE",
            "malzeme": "Beyaz Mdflam",
            "stok_kodu": "STK001",
            "bant_kalinligi": "0.40 MM",
            "grain_varsayilan": 3,
            "rows": [
                {
                    "satir_sirasi": 1,
                    "malzeme": "Beyaz Mdflam",
                    "boy": 500,
                    "en": 300,
                    "adet": 2,
                    "grain": 3,
                    "bilgi": "Kapak",
                    "u1": True,
                    "u2": False,
                    "k1": False,
                    "k2": False,
                    "delik_1": "123",
                    "delik_2": "456",
                    "satir_kaynagi": "OCR",
                },
                {
                    "satir_sirasi": 2,
                    "malzeme": "Beyaz Mdflam",
                    "boy": 450,
                    "en": 250,
                    "adet": 1,
                    "grain": 3,
                    "bilgi": "Yan Panel",
                    "u1": False,
                    "u2": False,
                    "k1": False,
                    "k2": False,
                    "delik_1": None,
                    "delik_2": None,
                    "satir_kaynagi": "MANUEL",
                },
            ],
        },
    )

    preview = optiplan_workflow_service.export_preview(db_session, record["kayit_uuid"], xlsx_aktif_mi=True)

    assert len(preview["satirlar"]) == 2
    assert preview["export_manifest"]["requested_formats"] == ["xlsx"]
    for row in preview["satirlar"]:
        assert list(row.keys()) == EXPORT_COLUMNS

    bandli = next(row for row in preview["satirlar"] if row["[P_LENGTH]"] == 500)
    bantsiz = next(row for row in preview["satirlar"] if row["[P_LENGTH]"] == 450)

    assert bandli["[P_EDGE_MAT_UP]"] == "04"
    assert bandli["[P_IIDESC]"] == "123"
    assert bandli["[P_DESC1]"] == "456"

    assert bantsiz["[P_EDGE_MAT_UP]"] == ""
    assert bantsiz["[P_EGDE_MAT_LO]"] == ""
    assert bantsiz["[P_IIDESC]"] == ""
    assert bantsiz["[P_DESC1]"] == ""


def test_export_preview_rejects_invalid_edge_code_from_contract_guard(
    db_session,
    workspace_tmp_path: Path,
    monkeypatch,
):
    _configure_settings(db_session, workspace_tmp_path)
    record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="invalid-edge.pdf",
        content=b"edge-verisi",
        kaynak_klasor="email_raw",
    )

    optiplan_workflow_service.update_phase3(
        db_session,
        record["kayit_uuid"],
        {
            "cari_unvan": "Ornek Cari",
            "cari_kodu": "CARI001",
            "siparis_no": "SIP-EDGE-001",
            "termin": "2026-03-12",
            "teslim_tarihi": "2026-03-13",
            "teslimat_adresi": "Edge Depo",
            "odeme_sekli": "HAVALE",
            "malzeme": "Beyaz Mdflam",
            "stok_kodu": "STK001",
            "bant_kalinligi": "0.40 MM",
            "rows": [
                {
                    "satir_sirasi": 1,
                    "malzeme": "Beyaz Mdflam",
                    "boy": 500,
                    "en": 300,
                    "adet": 1,
                    "grain": 3,
                    "u1": True,
                    "satir_kaynagi": "MANUEL",
                }
            ],
        },
    )

    monkeypatch.setattr(optiplan_workflow_service, "_edge_value", lambda enabled, bant_value: "ZZ")

    with pytest.raises(ValidationError):
        optiplan_workflow_service.export_preview(db_session, record["kayit_uuid"], xlsx_aktif_mi=True)


def test_export_preview_rejects_invalid_grain_value(db_session, workspace_tmp_path: Path):
    _configure_settings(db_session, workspace_tmp_path)
    record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="invalid-grain.pdf",
        content=b"grain-verisi",
        kaynak_klasor="email_raw",
    )

    optiplan_workflow_service.update_phase3(
        db_session,
        record["kayit_uuid"],
        {
            "cari_unvan": "Ornek Cari",
            "cari_kodu": "CARI001",
            "siparis_no": "SIP-GRAIN-001",
            "termin": "2026-03-12",
            "teslim_tarihi": "2026-03-13",
            "teslimat_adresi": "Grain Depo",
            "odeme_sekli": "HAVALE",
            "malzeme": "Beyaz Mdflam",
            "stok_kodu": "STK001",
            "rows": [
                {
                    "satir_sirasi": 1,
                    "malzeme": "Beyaz Mdflam",
                    "boy": 500,
                    "en": 300,
                    "adet": 1,
                    "grain": 7,
                    "satir_kaynagi": "MANUEL",
                }
            ],
        },
    )

    with pytest.raises(ValidationError):
        optiplan_workflow_service.export_preview(db_session, record["kayit_uuid"], xlsx_aktif_mi=True)


def test_phase3_rejects_invalid_satir_kaynagi(db_session, workspace_tmp_path: Path):
    _configure_settings(db_session, workspace_tmp_path)
    record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="invalid-source.pdf",
        content=b"siparis-verisi",
        kaynak_klasor="email_raw",
    )

    with pytest.raises(ValidationError):
        optiplan_workflow_service.update_phase3(
            db_session,
            record["kayit_uuid"],
            {
                "cari_unvan": "Ornek Cari",
                "cari_kodu": "CARI001",
                "siparis_no": "SIP-000002",
                "termin": "2026-03-12",
                "teslim_tarihi": "2026-03-13",
                "teslimat_adresi": "Invalid Source Depo",
                "odeme_sekli": "HAVALE",
                "malzeme": "Beyaz Mdflam",
                "stok_kodu": "STK001",
                "bant_kalinligi": "0.40 MM",
                "rows": [
                    {
                        "satir_sirasi": 1,
                        "malzeme": "Beyaz Mdflam",
                        "boy": 500,
                        "en": 300,
                        "adet": 1,
                        "grain": 3,
                        "satir_kaynagi": "API",
                    }
                ],
            },
        )


def test_phase3_accepts_camel_case_row_aliases(db_session, workspace_tmp_path: Path):
    _configure_settings(db_session, workspace_tmp_path)
    record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="camel-case.pdf",
        content=b"siparis-verisi",
        kaynak_klasor="email_raw",
    )

    updated = optiplan_workflow_service.update_phase3(
        db_session,
        record["kayit_uuid"],
        {
            "cari_unvan": "Ornek Cari",
            "cari_kodu": "CARI001",
            "siparis_no": "SIP-000003",
            "termin": "2026-03-12",
            "teslim_tarihi": "2026-03-13",
            "teslimat_adresi": "Camel Case Depo",
            "odeme_sekli": "HAVALE",
            "malzeme": "Beyaz Mdflam",
            "stok_kodu": "STK001",
            "bant_kalinligi": "0.40 MM",
            "rows": [
                {
                    "satirSirasi": 1,
                    "malzeme": "Beyaz Mdflam",
                    "boy": 500,
                    "en": 300,
                    "adet": 1,
                    "grain": 3,
                    "delik1": "12",
                    "delik2": "34",
                    "satirKaynagi": "MANUEL",
                    "plakaRef": "PLAKA-1",
                    "bantKalinligiOverride": "1 MM",
                }
            ],
        },
    )

    row = updated["satirlar"][0]
    assert row["satir_sirasi"] == 1
    assert row["delik_1"] == "12"
    assert row["delik_2"] == "34"
    assert row["satir_kaynagi"] == "MANUEL"
    assert row["plaka_ref"] == "PLAKA-1"
    assert row["bant_kalinligi_override"] == "1 MM"


def test_phase3_accepts_camel_case_plate_aliases(db_session, workspace_tmp_path: Path):
    _configure_settings(db_session, workspace_tmp_path)
    record = optiplan_workflow_service.manual_import(
        db_session,
        file_name="camel-plate.pdf",
        content=b"siparis-verisi",
        kaynak_klasor="email_raw",
    )

    updated = optiplan_workflow_service.update_phase3(
        db_session,
        record["kayit_uuid"],
        {
            "cari_unvan": "Ornek Cari",
            "cari_kodu": "CARI001",
            "siparis_no": "SIP-000004",
            "termin": "2026-03-12",
            "teslim_tarihi": "2026-03-13",
            "teslimat_adresi": "Camel Plate Depo",
            "odeme_sekli": "HAVALE",
            "malzeme": "Beyaz Mdflam",
            "stok_kodu": "STK001",
            "rows": [
                {
                    "satirSirasi": 1,
                    "malzeme": "Beyaz Mdflam",
                    "boy": 500,
                    "en": 300,
                    "adet": 1,
                    "grain": 3,
                    "satirKaynagi": "MANUEL",
                    "plakaRef": "PLAKA-2",
                }
            ],
            "plates": [
                {
                    "plakaRef": "PLAKA-2",
                    "etiket": "PLAKA-2 (2500x1850)",
                    "plakaBoyMm": 2500,
                    "plakaEnMm": 1850,
                    "genelListedeMi": True,
                }
            ],
        },
    )

    plate = next(item for item in updated["plakalar"] if item["plaka_ref"] == "PLAKA-2")
    assert plate["plaka_boy_mm"] == 2500
    assert plate["plaka_en_mm"] == 1850
    assert plate["genel_listede_mi"] is True


def test_mark_error_requires_note_and_retry_clones_record(db_session, workspace_tmp_path: Path):
    kayit_uuid, _ = _create_record_with_row(db_session, workspace_tmp_path)

    with pytest.raises(ValidationError):
        optiplan_workflow_service.mark_error(
            db_session,
            kayit_uuid,
            hata_fazi="PHASE_2",
            hata_nedeni="Diger",
            operator_notu=None,
        )

    errored = optiplan_workflow_service.mark_error(
        db_session,
        kayit_uuid,
        hata_fazi="PHASE_2",
        hata_nedeni="Diger",
        operator_notu="Operatör notu",
    )
    retry = optiplan_workflow_service.retry_record(db_session, kayit_uuid)

    assert errored["dosya_durumu"] == "HATALI"
    assert len(list(_case_dir(workspace_tmp_path, "hatali").iterdir())) == 1
    assert retry["retry_no"] == 1
    assert len(retry["satirlar"]) == 1
    assert len(retry["plakalar"]) == 1


def test_mikro_searches_return_mapped_items(monkeypatch):
    def fake_read_rows(query, params):
        if "CARI_HESAPLAR" in query:
            return [{"CARI_KOD": "C001", "CARI_UNVAN": "Deneme Cari", "TELEFON1": "555"}]
        return [{"STOK_KOD": "S001", "STOK_ISIM": "Deneme Stok"}]

    monkeypatch.setattr("app.services.optiplan_workflow_service._read_mikro_rows", fake_read_rows)

    assert optiplan_workflow_service.search_customers("deneme") == [
        {"Cari_Kodu": "C001", "Cari_Unvan": "Deneme Cari", "Telefon": "555"}
    ]
    assert optiplan_workflow_service.search_stocks("deneme") == [
        {"Stok_Kodu": "S001", "Stok_Adi": "Deneme Stok"}
    ]




