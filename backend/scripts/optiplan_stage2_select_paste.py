#!/usr/bin/env python3
"""OptiPlanning stage-2 automation.

Flow:
1) Read XLSX active sheet from row 1 and copy full used range as TSV.
2) Focus OptiPlanning main window.
3) Select outer tab: Parcalar.
4) Select inner tab: Sec.
5) Right-click in list area.
6) Click context menu item: Yapistir (fallback: Ctrl+V).
"""

from __future__ import annotations

import argparse
import time
from pathlib import Path

import win32clipboard
from openpyxl import load_workbook
from pywinauto import Desktop, mouse
from pywinauto.keyboard import send_keys
from win32com.client import DispatchEx

from optiplan_template_contract import validate_template_contract


def _normalize_text(value: str) -> str:
    s = (value or "").strip().lower()
    for a, b in [("\u00e7", "c"), ("\u015f", "s"), ("\u011f", "g"), ("\u00fc", "u"), ("\u00f6", "o"), ("\u0131", "i")]:
        s = s.replace(a, b)
    return s


def _cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _extract_full_sheet_tsv(xlsx_path: Path) -> tuple[str, int, int]:
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))

    last_row = 0
    last_col = 0
    for r_idx, row in enumerate(values, start=1):
        row_has_value = False
        for c_idx, val in enumerate(row, start=1):
            if val is None:
                continue
            if isinstance(val, str) and not val.strip():
                continue
            row_has_value = True
            if c_idx > last_col:
                last_col = c_idx
        if row_has_value:
            last_row = r_idx

    if last_row == 0 or last_col == 0:
        raise ValueError("XLSX icinde kopyalanacak dolu hucre bulunamadi")

    lines: list[str] = []
    for r_idx in range(1, last_row + 1):
        row_cells = [_cell_to_text(ws.cell(row=r_idx, column=c_idx).value) for c_idx in range(1, last_col + 1)]
        lines.append("\t".join(row_cells))
    return ("\n".join(lines), last_row, last_col)


def _copy_from_excel_clipboard(xlsx_path: Path, sheet_name: str | None = None) -> tuple[int, int]:
    excel = DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(str(xlsx_path), ReadOnly=True)
        ws = wb.Worksheets(sheet_name) if sheet_name else wb.ActiveSheet
        used = ws.UsedRange
        last_row = int(used.Row + used.Rows.Count - 1)
        last_col = int(used.Column + used.Columns.Count - 1)
        if last_row < 1 or last_col < 1:
            raise ValueError("Excel used range bos gorunuyor")
        ws.Range(ws.Cells(1, 1), ws.Cells(last_row, last_col)).Copy()
        # Keep clipboard content and close workbook without save.
        wb.Close(SaveChanges=False)
        wb = None
        excel.Quit()
        return (last_row, last_col)
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass


def _set_clipboard_text(text: str) -> None:
    win32clipboard.OpenClipboard()
    try:
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardText(text)
    finally:
        win32clipboard.CloseClipboard()


def _get_clipboard_text() -> str:
    win32clipboard.OpenClipboard()
    try:
        data = win32clipboard.GetClipboardData()
        return str(data)
    except Exception:
        return ""
    finally:
        win32clipboard.CloseClipboard()


def _find_main_uia():
    cands = []
    for w in Desktop(backend="uia").windows():
        try:
            if not w.is_visible():
                continue
            title = (w.window_text() or "").lower()
            if "optiplanning - [" in title or "professional (hp)" in title:
                cands.append(w)
        except Exception:
            continue
    return cands[-1] if cands else None


def _find_main_win32():
    cands = []
    for w in Desktop(backend="win32").windows():
        try:
            if not w.is_visible():
                continue
            title = (w.window_text() or "").lower()
            if "optiplanning - [" in title or "professional (hp)" in title:
                cands.append(w)
        except Exception:
            continue
    return cands[-1] if cands else None


def _find_tab_item(main_uia, wanted_norm: str, outer: bool):
    for ti in main_uia.descendants(control_type="TabItem"):
        try:
            if not ti.is_visible():
                continue
            name = _normalize_text(ti.window_text() or "")
            if name != wanted_norm:
                continue
            if outer and ti.rectangle().top > 110:
                continue
            if (not outer) and ti.rectangle().top <= 110:
                continue
            return ti
        except Exception:
            continue
    return None


def _select_tab(tab_item) -> None:
    try:
        tab_item.select()
    except Exception:
        tab_item.click_input()
    time.sleep(0.25)


def _click_context_menu_item(tokens: list[str], debug: bool = False) -> bool:
    tokens = [_normalize_text(t) for t in tokens]
    for w in Desktop(backend="uia").windows():
        try:
            if not w.is_visible():
                continue
            for mi in w.descendants(control_type="MenuItem"):
                if not mi.is_visible():
                    continue
                nm = _normalize_text(mi.window_text() or "")
                if debug and any(tok in nm for tok in tokens):
                    print(f"[DEBUG] popup candidate: {mi.window_text()!r}")
                if any(tok in nm for tok in tokens):
                    mi.click_input()
                    if debug:
                        print(f"[DEBUG] clicked menu: {mi.window_text()!r}")
                    return True
        except Exception:
            continue
    return False


def _find_primary_list_target(main_win32):
    best = None
    best_area = 0
    try:
        for c in main_win32.descendants():
            try:
                if not c.is_visible():
                    continue
                cls = c.class_name()
                if cls not in ("SysListView32", "ListBox"):
                    continue
                r = c.rectangle()
                w = max(0, int(r.right - r.left))
                h = max(0, int(r.bottom - r.top))
                area = w * h
                if area > best_area:
                    best_area = area
                    best = (r.left + min(120, max(20, w // 4)), r.top + min(120, max(20, h // 4)))
            except Exception:
                continue
    except Exception:
        return None
    return best


def _find_primary_list_control(main_win32):
    best = None
    best_area = 0
    try:
        for c in main_win32.descendants():
            try:
                if not c.is_visible():
                    continue
                cls = c.class_name()
                if cls not in ("SysListView32", "ListBox"):
                    continue
                r = c.rectangle()
                w = max(0, int(r.right - r.left))
                h = max(0, int(r.bottom - r.top))
                area = w * h
                if area > best_area:
                    best_area = area
                    best = c
            except Exception:
                continue
    except Exception:
        return None
    return best


def _control_item_count(ctrl) -> int | None:
    if ctrl is None:
        return None
    try:
        cls = ctrl.class_name()
    except Exception:
        return None
    try:
        if cls == "SysListView32":
            return int(ctrl.item_count())
        if cls == "ListBox":
            return int(len(ctrl.item_texts()))
    except Exception:
        return None
    return None


def _wait_item_count_increase(ctrl, before: int | None, timeout_sec: float = 4.0) -> bool:
    if before is None:
        return False
    end = time.time() + timeout_sec
    while time.time() < end:
        cur = _control_item_count(ctrl)
        if cur is not None and cur > before:
            return True
        time.sleep(0.2)
    return False


def _verify_paste_result(target: tuple[int, int], debug: bool = False) -> bool:
    sentinel = "__OPTI_VERIFY_SENTINEL__"
    _set_clipboard_text(sentinel)

    # Keyboard-first verification (usually more reliable on grid controls).
    mouse.click(button="left", coords=target)
    time.sleep(0.1)
    send_keys("^a")
    time.sleep(0.05)
    send_keys("^c")
    time.sleep(0.2)
    copied = _get_clipboard_text()
    if debug:
        print(f"[DEBUG] verify-kbd clipboard_len={len(copied)}")
    if copied.strip() and copied != sentinel:
        return True

    # Context-menu fallback verification.
    _set_clipboard_text(sentinel)
    mouse.click(button="right", coords=target)
    time.sleep(0.2)
    _click_context_menu_item(["tumunu sec", "select all"], debug=debug)
    time.sleep(0.15)
    mouse.click(button="right", coords=target)
    time.sleep(0.2)
    copied_ok = _click_context_menu_item(["kopyala", "copy"], debug=debug)
    if not copied_ok:
        send_keys("^c")
    time.sleep(0.2)
    copied = _get_clipboard_text()
    if debug:
        print(f"[DEBUG] verify-menu copy_clicked={copied_ok} clipboard_len={len(copied)}")
    return bool(copied.strip() and copied != sentinel)


def run_stage2(
    xlsx_path: Path,
    timeout_sec: float = 20.0,
    debug: bool = False,
    copy_mode: str = "excel",
) -> tuple[int, int]:
    contract = validate_template_contract(xlsx_path)
    if not contract.ok:
        raise ValueError(contract.message + " | " + " ; ".join(contract.mismatches))

    copied_rows = 0
    copied_cols = 0
    if copy_mode == "excel":
        try:
            copied_rows, copied_cols = _copy_from_excel_clipboard(xlsx_path)
        except Exception as exc:
            if debug:
                print(f"[DEBUG] excel copy failed, fallback tsv: {exc}")
            tsv, copied_rows, copied_cols = _extract_full_sheet_tsv(xlsx_path)
            _set_clipboard_text(tsv)
    else:
        tsv, copied_rows, copied_cols = _extract_full_sheet_tsv(xlsx_path)
        _set_clipboard_text(tsv)

    main = _find_main_uia()
    if not main:
        raise RuntimeError("OptiPlanning ana pencere bulunamadi")

    try:
        main.set_focus()
    except Exception:
        pass

    end = time.time() + timeout_sec
    parca_tab = None
    sec_tab = None
    while time.time() < end:
        parca_tab = _find_tab_item(main, "parcalar", outer=True)
        sec_tab = _find_tab_item(main, "sec", outer=False)
        if parca_tab and sec_tab:
            break
        time.sleep(0.2)

    if not parca_tab:
        raise RuntimeError("'Parcalar' sekmesi bulunamadi")
    if not sec_tab:
        raise RuntimeError("'Sec' sekmesi bulunamadi")

    if debug:
        print(f"[DEBUG] outer tab: {parca_tab.window_text()!r}")
        print(f"[DEBUG] inner tab: {sec_tab.window_text()!r}")
        print(f"[DEBUG] clipboard range: {copied_rows}x{copied_cols}")

    _select_tab(parca_tab)
    _select_tab(sec_tab)

    target = None
    list_ctrl = None
    before_count = None
    main32 = _find_main_win32()
    if main32 is not None:
        target = _find_primary_list_target(main32)
        list_ctrl = _find_primary_list_control(main32)
        before_count = _control_item_count(list_ctrl)
    if target is None:
        srect = sec_tab.rectangle()
        target = (srect.left + 180, srect.bottom + 180)
    if debug:
        print(f"[DEBUG] paste target: {target}")
        print(f"[DEBUG] item_count_before: {before_count}")

    mouse.click(button="left", coords=target)
    time.sleep(0.1)
    mouse.click(button="right", coords=target)
    time.sleep(0.25)

    pasted = _click_context_menu_item(["yapistir", "paste"], debug=debug)
    if not pasted:
        if debug:
            print("[DEBUG] menu item not found, fallback Ctrl+V")
        mouse.click(button="left", coords=target)
        time.sleep(0.1)
        send_keys("^v")
    time.sleep(0.25)

    if _wait_item_count_increase(list_ctrl, before_count, timeout_sec=min(6.0, timeout_sec)):
        if debug:
            after_count = _control_item_count(list_ctrl)
            print(f"[DEBUG] item_count_after: {after_count}")
        return (copied_rows, copied_cols)

    if not _verify_paste_result(target, debug=debug):
        # Retry once with direct keyboard paste, then verify again.
        if debug:
            print("[DEBUG] first verification failed, retrying paste with Ctrl+V")
        mouse.click(button="left", coords=target)
        time.sleep(0.1)
        send_keys("^v")
        time.sleep(0.3)
        if _wait_item_count_increase(list_ctrl, before_count, timeout_sec=min(6.0, timeout_sec)):
            if debug:
                after_count = _control_item_count(list_ctrl)
                print(f"[DEBUG] item_count_after_retry: {after_count}")
            return (copied_rows, copied_cols)
        if not _verify_paste_result(target, debug=debug):
            raise RuntimeError("Yapistirma dogrulamasi basarisiz (secili veri kopyalanamadi)")

    return (copied_rows, copied_cols)


def parse_args():
    p = argparse.ArgumentParser(description="OptiPlanning stage-2: XLSX full range copy -> Sec -> Yapistir")
    p.add_argument("--xlsx", required=True, help="Yapistirilacak xlsx dosyasi")
    p.add_argument("--timeout", type=float, default=20.0, help="Kontrol bekleme suresi (sn)")
    p.add_argument("--debug", action="store_true", help="Debug loglarini yazdir")
    p.add_argument(
        "--copy-mode",
        choices=["excel", "tsv"],
        default="excel",
        help="Pano kopyalama modu (varsayilan: excel)",
    )
    return p.parse_args()


def main():
    args = parse_args()
    path = Path(args.xlsx)
    if not path.exists():
        raise SystemExit(f"HATA: XLSX bulunamadi: {path}")

    try:
        rows, cols = run_stage2(path, timeout_sec=args.timeout, debug=args.debug, copy_mode=args.copy_mode)
        print(f"TAMAM: Stage-2 uygulandi. Kopyalanan aralik: {rows} satir x {cols} kolon")
        return 0
    except Exception as exc:
        print(f"HATA: Stage-2 basarisiz: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
