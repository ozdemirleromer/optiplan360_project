#!/usr/bin/env python3
"""
Template contract for OptiPlanning XLSX import sheets.

The first 2 rows are treated as a strict schema contract.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_ROW1 = [
    "[P_CODE_MAT]",
    "[P_LENGTH]",
    "[P_WIDTH]",
    "[P_MINQ]",
    "[P_GRAIN]",
    "[P_IDESC]",
    "[P_EDGE_MAT_UP]",
    "[P_EGDE_MAT_LO]",
    "[P_EDGE_MAT_SX]",
    "[P_EDGE_MAT_DX]",
    "[P_IIDESC]",
    "[P_DESC1]",
]

EXPECTED_ROW2 = [
    "Material",
    "Length",
    "Width",
    "Min Q.",
    "GrainI",
    "Description",
    "Upper strip mat.",
    "Lower strip mat.",
    "Left strip mat.",
    "Right strip mat.",
    "II Description",
    "Description 1",
]


@dataclass
class ContractResult:
    ok: bool
    message: str
    column_count: int
    mismatches: list[str]


def _as_text(v) -> str:
    return "" if v is None else str(v).strip()


def validate_template_contract(
    xlsx_path: Path,
    sheet_name: str | None = None,
) -> ContractResult:
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    row1 = [_as_text(ws.cell(row=1, column=i).value) for i in range(1, len(EXPECTED_ROW1) + 1)]
    row2 = [_as_text(ws.cell(row=2, column=i).value) for i in range(1, len(EXPECTED_ROW2) + 1)]

    mismatches: list[str] = []
    for i, (exp, got) in enumerate(zip(EXPECTED_ROW1, row1), start=1):
        if exp != got:
            mismatches.append(f"R1C{i}: beklenen='{exp}' gelen='{got}'")
    for i, (exp, got) in enumerate(zip(EXPECTED_ROW2, row2), start=1):
        if exp != got:
            mismatches.append(f"R2C{i}: beklenen='{exp}' gelen='{got}'")

    if mismatches:
        return ContractResult(
            ok=False,
            message="XLSX 1-2. satir import sozlesmesine uymuyor",
            column_count=len(EXPECTED_ROW1),
            mismatches=mismatches,
        )

    return ContractResult(
        ok=True,
        message="XLSX 1-2. satir import sozlesmesi dogrulandi",
        column_count=len(EXPECTED_ROW1),
        mismatches=[],
    )

