"""
OptiPlan 360 — Order Router
Sipariş CRUD + Validasyon + Onay + Export
"""

import io
import json
import logging
import os
import re
import unicodedata
from datetime import datetime, timezone
from typing import List
from uuid import uuid4

from app.auth import get_current_user, require_operator, require_permissions
from app.database import get_db
from app.exceptions import BusinessRuleError, NotFoundError, ValidationError
from app.models import Customer, Order, OrderPart, User
from app.permissions import Permission
from app.schemas import (
    ExportFile,
    ExportResult,
    OrderCreate,
    OrderListResponse,
    OrderOut,
    OrderPartCreate,
    OrderPartOut,
    OrderUpdate,
    ValidationResult,
)
from app.services.optimization import get_export_rows
from app.services.order_service import OrderService
from app.utils import create_audit_log
from fastapi import APIRouter, Depends, File, Form, Query, UploadFile
from fastapi.responses import FileResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session, joinedload, subqueryload

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/orders", tags=["orders"])

EXPORT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)

EXPORT_BASE_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "exports")
os.makedirs(EXPORT_BASE_DIR, exist_ok=True)


# ─── Sipariş Listesi ───
@router.get("", response_model=OrderListResponse)
def list_orders(
    status: str = Query(None),
    customer_id: str = Query(None),
    search: str = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _: User = Depends(require_permissions(Permission.ORDERS_VIEW)),
):
    q = db.query(Order).options(subqueryload(Order.parts), joinedload(Order.customer))

    if status:
        q = q.filter(Order.status == status)
    if customer_id:
        q = q.filter(Order.customer_id == customer_id)
    if search:
        s = f"%{search}%"
        q = q.filter(
            (Order.crm_name_snapshot.ilike(s))
            | (Order.material_name.ilike(s))
            | (Order.phone_norm.ilike(s))
            | (Order.id.ilike(s))
        )

    total = q.count()
    orders = q.order_by(Order.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()

    return OrderListResponse(
        data=[OrderService.order_to_list_item(o) for o in orders],
        total=total,
        page=page,
    )


# ─── Dosya İndirme (/{order_id}'den ÖNCE tanımlanmalı) ───
@router.get("/export/download/{filename}")
def download_export_file(
    filename: str,
    _=Depends(get_current_user),
):
    """Export edilmiş XLSX dosyasını indir"""
    # Güvenlik: path traversal koruması
    safe_name = os.path.basename(filename)
    if safe_name != filename or ".." in filename:
        raise BusinessRuleError("Geçersiz dosya adı")

    file_path = os.path.join(EXPORT_DIR, safe_name)
    if not os.path.isfile(file_path):
        raise NotFoundError("Dosya")

    return FileResponse(
        path=file_path,
        filename=safe_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ─── Tek Sipariş ───
@router.get("/{order_id}", response_model=OrderOut)
def get_order(
    order_id: str,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    # ID validasyonu: integer değilse veya negatifse 404 dön
    try:
        order_id_int = int(order_id)
        if order_id_int < 0:
            raise NotFoundError("Sipariş")
    except ValueError:
        raise NotFoundError("Kaynak")

    order = (
        db.query(Order)
        .options(joinedload(Order.parts), joinedload(Order.audit_logs))
        .filter(Order.id == order_id_int)
        .first()
    )
    if not order:
        raise NotFoundError("Sipariş")
    return _order_to_out(order)


# ─── Sipariş Oluştur ───
@router.post("", response_model=OrderOut, status_code=201)
def create_order(
    body: OrderCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_operator),
):
    order = OrderService.create_order(db, body, user)
    return OrderService.order_to_out(order)


# ─── Parça Ekle (Mevcut Siparişe) ───
@router.post("/{order_id}/parts", response_model=OrderOut)
def add_parts(
    order_id: str,
    parts: list[OrderPartCreate],
    db: Session = Depends(get_db),
    user: User = Depends(require_operator),
):
    order = OrderService.add_parts(db, order_id, parts, user)
    return OrderService.order_to_out(order)


# ─── Validasyon ───
@router.post("/{order_id}/validate", response_model=ValidationResult)
def validate_order(
    order_id: str,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return OrderService.validate_order(db, order_id)


# ─── Onay ───
@router.post("/{order_id}/approve", response_model=OrderOut)
def approve_order(
    order_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(require_operator),
):
    order = OrderService.approve_order(db, order_id, user)
    return OrderService.order_to_out(order)


# ─── Durum Geçiş Kuralları ───
VALID_TRANSITIONS = {
    "NEW": ["HOLD", "CANCELLED", "IN_PRODUCTION"],
    "HOLD": ["NEW", "CANCELLED", "IN_PRODUCTION"],
    "IN_PRODUCTION": ["HOLD", "CANCELLED", "READY"],
    "READY": ["IN_PRODUCTION", "DELIVERED"],
    "DELIVERED": ["DONE"],
    "DONE": [],  # Son durum — geçiş yok
    "CANCELLED": ["NEW"],  # İptal → tekrar yeni yapılabilir
}


# ─── Geçerli Geçişleri Getir ───
@router.get("/{order_id}/transitions")
def get_transitions(
    order_id: str,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    return OrderService.get_transitions(db, order_id)


# ─── Durum Güncelle ───
@router.patch("/{order_id}/status")
async def update_status(
    order_id: str,
    new_status: str = Query(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_operator),
):
    return OrderService.update_status(db, order_id, new_status, user)


# ─── OptiPlanning Export ───
def _load_share_path() -> str | None:
    """system_config.json'dan makine paylaşım yolunu oku"""
    config_path = os.path.join(
        os.path.dirname(__file__), "..", "..", "..", "config", "system_config.json"
    )
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        return cfg.get("share_path_machine") or None
    except Exception:
        return None


def _safe_filename(name: str) -> str:
    """Dosya adı için güvenli karakter dönüşümü"""
    safe = name.replace(" ", "_").upper()
    safe = re.sub(r'[<>:"/\\|?*]', "_", safe)
    return safe


@router.post("/{order_id}/export/opti", response_model=ExportResult)
def export_opti(
    order_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(require_operator),
):
    validate_result = validate_order(order_id, db, user)
    if not validate_result.valid:
        raise BusinessRuleError("Export başarısız: Validasyon hataları mevcut")

    order = db.query(Order).options(joinedload(Order.parts)).filter(Order.id == order_id).first()
    if not order:
        raise NotFoundError("Sipariş")
    if order.status not in ("IN_PRODUCTION", "NEW", "READY"):
        raise BusinessRuleError("Sipariş uygun durumda değil")

    files: list[ExportFile] = []

    # Export dizinini oluştur
    export_dir = os.path.join(EXPORT_DIR, f"{order.crm_name_snapshot}_{order.ts_code}")
    os.makedirs(export_dir, exist_ok=True)

    # Gövde ve arkalık parçalarını renk ve kalınlık bazında ayır
    grouped_parts = {}
    for part in order.parts:
        key = (part.part_group, part.color, part.thickness_mm)
        if key not in grouped_parts:
            grouped_parts[key] = []
        grouped_parts[key].append(part)

    name_safe = _safe_filename(order.crm_name_snapshot)

    for (part_group, color, thickness), parts in grouped_parts.items():
        group_name = "GOVDE" if part_group == "GOVDE" else "ARKALIK"
        fn = f"{name_safe}_{order.ts_code}_{thickness}mm_{color}_{group_name}.xlsx"
        path = os.path.join(export_dir, fn)
        rows = get_export_rows(parts, part_group=part_group)
        _create_opti_xlsx(path, rows, order, group_name)
        download_url = f"/api/v1/orders/export/download/{fn}"
        files.append(
            ExportFile(filename=fn, part_group=group_name, path=path, download_url=download_url)
        )

    return ExportResult(files=files)


# ─── Yardımcılar ───
def _validate_part(p: OrderPartCreate):
    """Arkalıkta bant kontrolü (Handoff §0.4)"""
    if p.part_group == "ARKALIK" and any([p.u1, p.u2, p.k1, p.k2]):
        raise BusinessRuleError("Arkalıkta bant olamaz (Handoff §0.4)")


def _order_to_out(order: Order) -> OrderOut:
    """Geriye uyumluluk wrapper'ı — OrderService.order_to_out() kullanır."""
    return OrderService.order_to_out(order)


def _create_opti_xlsx(path: str, rows: list, order: Order, part_group: str = "GOVDE"):
    """OptiPlanning uyumlu xlsx oluştur — metadata + parça tablosu. rows: get_export_rows() çıktısı (Grain §0.3 uygulanmış)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "OptiPlan"

    # Stiller
    label_font = Font(bold=True, size=10, color="1F4E79")
    label_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    value_font = Font(size=10)
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ─── Satır 1: Sipariş bilgileri ───
    meta_labels_1 = [
        ("Sipariş", order.ts_code),
        ("Müşteri", order.crm_name_snapshot),
        ("Tarih", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")),
    ]
    col = 1
    for label, value in meta_labels_1:
        c = ws.cell(row=1, column=col, value=label)
        c.font = label_font
        c.fill = label_fill
        c.border = thin_border
        c = ws.cell(row=1, column=col + 1, value=value)
        c.font = value_font
        c.border = thin_border
        col += 2

    # ─── Satır 2: Malzeme / Plaka bilgileri ───
    thickness = order.thickness_mm if part_group == "GOVDE" else 5
    band_display = f"{order.band_mm} mm" if order.band_mm and part_group == "GOVDE" else "-"
    meta_labels_2 = [
        ("Malzeme", order.material_name),
        ("Kalınlık", f"{thickness} mm"),
        ("Plaka", f"{order.plate_w_mm} x {order.plate_h_mm}"),
        ("Renk", order.color),
        ("Bant", band_display),
        ("Grup", part_group),
    ]
    col = 1
    for label, value in meta_labels_2:
        c = ws.cell(row=2, column=col, value=label)
        c.font = label_font
        c.fill = label_fill
        c.border = thin_border
        c = ws.cell(row=2, column=col + 1, value=value)
        c.font = value_font
        c.border = thin_border
        col += 2

    # ─── Satır 3: Boş ayırıcı ───

    # ─── Satır 4: Tablo header ───
    headers = ["Boy", "En", "Adet", "Grain", "U1", "U2", "K1", "K2", "Açıklama", "Delik1", "Delik2"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    # ─── Satır 5+: Parça verileri (ExportRow — grain §0.3 normalize, @437/@433 metadata) ───
    for row_idx, p in enumerate(rows, 5):
        ws.cell(row=row_idx, column=1, value=p.boy_mm).border = thin_border
        ws.cell(row=row_idx, column=2, value=p.en_mm).border = thin_border
        ws.cell(row=row_idx, column=3, value=p.adet).border = thin_border
        ws.cell(row=row_idx, column=4, value=p.grain_code).border = thin_border
        ws.cell(row=row_idx, column=5, value=1 if p.u1 else 0).border = thin_border
        ws.cell(row=row_idx, column=6, value=1 if p.u2 else 0).border = thin_border
        ws.cell(row=row_idx, column=7, value=1 if p.k1 else 0).border = thin_border
        ws.cell(row=row_idx, column=8, value=1 if p.k2 else 0).border = thin_border
        ws.cell(row=row_idx, column=9, value=(p.part_desc or "")).border = thin_border
        ws.cell(row=row_idx, column=10, value=(p.drill_code_1 or "")).border = thin_border
        ws.cell(row=row_idx, column=11, value=(p.drill_code_2 or "")).border = thin_border

    # Kolon genişlikleri
    widths = [12, 12, 8, 14, 5, 5, 5, 5, 20, 12, 12]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w
    # L kolonu (metadata overflow)
    ws.column_dimensions["L"].width = 14

    wb.save(path)


# ═══════════════════════════════════════════════════
# XLSX IMPORT - Excel dosyasından parça satırlarını içeri al
# ═══════════════════════════════════════════════════
def _normalize_text(value) -> str:
    s = str(value or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s


def _to_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_int(value):
    fv = _to_float(value)
    if fv is None:
        return None
    try:
        return int(fv)
    except (ValueError, TypeError):
        return None


def _parse_excel_bool(value) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    s = _normalize_text(value)
    if not s:
        return False
    return s not in {"0", "0.0", "false", "f", "no", "n", "yok", "hayir"}


def _parse_grain(value) -> str:
    if value is None:
        return "0-Material"
    s = str(value).strip()
    if not s:
        return "0-Material"
    if s in {"0-Material", "1-Material", "2-Material", "3-Material"}:
        return s
    n = _to_int(s)
    if n in (0, 1, 2, 3):
        return f"{n}-Material"
    m = re.match(r"^([0-3])\s*[-_ ]", s)
    if m:
        return f"{m.group(1)}-Material"
    return "0-Material"


def _header_key(text: str) -> str | None:
    t = _normalize_text(text)
    t = re.sub(r"[^a-z0-9]", "", t)
    mapping = {
        "boy": "boy",
        "length": "boy",
        "plength": "boy",
        "en": "en",
        "width": "en",
        "pwidth": "en",
        "adet": "adet",
        "qty": "adet",
        "quantity": "adet",
        "count": "adet",
        "minq": "adet",
        "pminq": "adet",
        "grain": "grain",
        "graini": "grain",
        "pgrain": "grain",
        "pgraini": "grain",
        "u1": "u1",
        "upperstripmat": "u1",
        "pedgematup": "u1",
        "u2": "u2",
        "lowerstripmat": "u2",
        "pedgematlo": "u2",
        "pegdematlo": "u2",
        "k1": "k1",
        "leftstripmat": "k1",
        "pedgematsx": "k1",
        "k2": "k2",
        "rightstripmat": "k2",
        "pedgematdx": "k2",
        "aciklama": "desc",
        "description": "desc",
        "pidesc": "desc",
        "info": "desc",
        "bilgi": "desc",
        "piidesc": "drill1",
        "iidescription": "drill1",
        "delik1": "drill1",
        "pdesc1": "drill2",
        "description1": "drill2",
        "delik2": "drill2",
    }
    return mapping.get(t)


def _extract_meta_and_rows(ws) -> tuple[list[dict], list[str], dict]:
    warnings: list[str] = []
    meta = {
        "thickness_mm": 18,
        "plate_w_mm": 2100,
        "plate_h_mm": 2800,
        "material_name": "Belirtilmedi",
    }

    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return [], ["Dosya bos."], meta

    header_idx = -1
    first_data_idx = -1
    inferred_offset = 0
    header_map: dict[str, int] = {}

    for i, row in enumerate(rows[:120]):
        cells = [str(c or "") for c in row]
        if not any(c.strip() for c in cells):
            continue

        for c in cells[:3]:
            u = str(c or "").upper()
            tm = re.search(r"(\d+)\s*MM", u)
            if tm:
                meta["thickness_mm"] = int(tm.group(1))
            pm = re.search(r"(\d+)\s*[*X]\s*(\d+)", u)
            if pm:
                w = int(pm.group(1))
                h = int(pm.group(2))
                meta["plate_w_mm"] = w * 10 if w < 1000 else w
                meta["plate_h_mm"] = h * 10 if h < 1000 else h
        for c in cells:
            cu = str(c or "")
            nu = _normalize_text(cu)
            if nu.startswith("renk") or nu.startswith("malzeme"):
                parts = re.split(r"[:=]", cu, maxsplit=1)
                if len(parts) > 1 and parts[1].strip():
                    meta["material_name"] = parts[1].strip()

        if header_idx == -1:
            local_map: dict[str, int] = {}
            for col, c in enumerate(cells):
                key = _header_key(c)
                if key and key not in local_map:
                    local_map[key] = col
            if local_map:
                header_idx = i
                header_map = local_map

        if first_data_idx == -1:
            has_standard = (
                _to_float(cells[0]) is not None
                and _to_float(cells[1]) is not None
                and _to_int(cells[2]) is not None
            )
            has_offset = (
                len(cells) > 3
                and _to_float(cells[1]) is not None
                and _to_float(cells[2]) is not None
                and _to_int(cells[3]) is not None
            )
            if has_standard or has_offset:
                first_data_idx = i
                inferred_offset = 1 if has_offset and not has_standard else 0

    if header_idx == -1 and first_data_idx == -1:
        return [], ["Dosya icinde veri satiri bulunamadi."], meta

    if header_idx != -1:
        idx = {
            "boy": header_map.get("boy"),
            "en": header_map.get("en"),
            "adet": header_map.get("adet"),
            "grain": header_map.get("grain"),
            "u1": header_map.get("u1"),
            "u2": header_map.get("u2"),
            "k1": header_map.get("k1"),
            "k2": header_map.get("k2"),
            "desc": header_map.get("desc"),
            "drill1": header_map.get("drill1"),
            "drill2": header_map.get("drill2"),
        }
        start_idx = header_idx + 1
        if idx["boy"] is None or idx["en"] is None or idx["adet"] is None:
            inferred_offset = (
                1
                if first_data_idx != -1
                and len(rows[first_data_idx]) > 3
                and _to_float(rows[first_data_idx][1]) is not None
                else 0
            )
            idx["boy"] = 1 if inferred_offset else 0
            idx["en"] = 2 if inferred_offset else 1
            idx["adet"] = 3 if inferred_offset else 2
            idx["grain"] = (
                idx["grain"] if idx["grain"] is not None else (4 if inferred_offset else 3)
            )
            idx["u1"] = idx["u1"] if idx["u1"] is not None else (5 if inferred_offset else 4)
            idx["u2"] = idx["u2"] if idx["u2"] is not None else (6 if inferred_offset else 5)
            idx["k1"] = idx["k1"] if idx["k1"] is not None else (7 if inferred_offset else 6)
            idx["k2"] = idx["k2"] if idx["k2"] is not None else (8 if inferred_offset else 7)
            idx["desc"] = (
                idx["desc"] if idx["desc"] is not None else (11 if inferred_offset else 10)
            )
            idx["drill1"] = (
                idx["drill1"] if idx["drill1"] is not None else (9 if inferred_offset else 8)
            )
            idx["drill2"] = (
                idx["drill2"] if idx["drill2"] is not None else (10 if inferred_offset else 9)
            )
    else:
        start_idx = first_data_idx
        idx = {
            "boy": 1 if inferred_offset else 0,
            "en": 2 if inferred_offset else 1,
            "adet": 3 if inferred_offset else 2,
            "grain": 4 if inferred_offset else 3,
            "u1": 5 if inferred_offset else 4,
            "u2": 6 if inferred_offset else 5,
            "k1": 7 if inferred_offset else 6,
            "k2": 8 if inferred_offset else 7,
            "drill1": 9 if inferred_offset else 8,
            "drill2": 10 if inferred_offset else 9,
            "desc": 11 if inferred_offset else 10,
        }

    parsed_parts: list[dict] = []
    for rix, row in enumerate(rows[start_idx:], start=start_idx + 1):
        if not row:
            continue

        def val(key):
            cidx = idx.get(key)
            if cidx is None or cidx < 0 or cidx >= len(row):
                return None
            return row[cidx]

        boy = _to_float(val("boy"))
        en = _to_float(val("en"))
        adet = _to_int(val("adet")) or 1

        if boy is None or en is None:
            continue
        if boy <= 0 or en <= 0:
            warnings.append(f"Satir {rix}: boy/en gecersiz, atlandi")
            continue
        if boy > 5000 or en > 5000:
            warnings.append(f"Satir {rix}: boy/en 5000mm'den buyuk, atlandi")
            continue
        if adet <= 0 or adet > 9999:
            warnings.append(f"Satir {rix}: adet gecersiz, atlandi")
            continue

        parsed_parts.append(
            {
                "boy_mm": boy,
                "en_mm": en,
                "adet": adet,
                "grain_code": _parse_grain(val("grain")),
                "u1": _parse_excel_bool(val("u1")),
                "u2": _parse_excel_bool(val("u2")),
                "k1": _parse_excel_bool(val("k1")),
                "k2": _parse_excel_bool(val("k2")),
                "part_desc": str(val("desc")).strip() if val("desc") not in (None, "") else None,
                "drill_code_1": (
                    str(val("drill1")).strip() if val("drill1") not in (None, "") else None
                ),
                "drill_code_2": (
                    str(val("drill2")).strip() if val("drill2") not in (None, "") else None
                ),
            }
        )

    return parsed_parts, warnings, meta


@router.post("/auto-create-from-xlsx", status_code=201)
async def auto_create_from_xlsx(
    list_name: str = Form(...),
    file: UploadFile = File(...),
    customer_phone: str = Form("5550000000"),
    part_group: str = Form("GOVDE"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_operator),
):
    """Yeni + Kaydet + ayni isim + XLSX olcu import adimlarini tek endpointte yapar."""
    raw_name = (list_name or "").strip()
    if not raw_name:
        raise ValidationError("Liste adi bos olamaz")
    if part_group not in ("GOVDE", "ARKALIK"):
        raise BusinessRuleError("part_group GOVDE veya ARKALIK olmali")

    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise ValidationError("Sadece .xlsx/.xls dosyasi desteklenir")
    file_stem = os.path.splitext(os.path.basename(filename))[0].strip()
    if _normalize_text(file_stem) != _normalize_text(raw_name):
        raise BusinessRuleError("Liste adi ile dosya adi (uzantisiz) ayni olmalidir")

    phone_digits = re.sub(r"\D+", "", customer_phone or "")
    if len(phone_digits) < 10:
        raise ValidationError("customer_phone en az 10 haneli olmalidir")

    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise ValidationError("Dosya boyutu 10MB'i asamaz")

    try:
        wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as exc:
        raise BusinessRuleError(f"Excel okunamadi: {exc}")

    parsed_parts, warnings, meta = _extract_meta_and_rows(ws)
    if not parsed_parts:
        raise BusinessRuleError("Dosyadan gecerli olcu satiri okunamadi")

    customer = db.query(Customer).filter(Customer.phone == phone_digits).first()
    if not customer:
        customer = Customer(
            name=raw_name,
            phone=phone_digits,
            created_by=current_user.id,
            updated_by=current_user.id,
        )
        db.add(customer)
        db.flush()

    ts_code = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    from sqlalchemy import func as sa_func

    max_no = db.query(sa_func.max(Order.order_no)).scalar() or 0

    order = Order(
        order_no=max_no + 1,
        customer_id=customer.id,
        crm_name_snapshot=raw_name,
        ts_code=ts_code,
        tracking_token=str(uuid4()),
        phone_norm=phone_digits,
        thickness_mm=meta["thickness_mm"],
        plate_w_mm=meta["plate_w_mm"],
        plate_h_mm=meta["plate_h_mm"],
        color=meta["material_name"],
        material_name=meta["material_name"],
        grain_default="0-Material",
        created_by=current_user.id,
        status="NEW",
    )
    db.add(order)
    db.flush()

    imported_rows = 0
    for p in parsed_parts:
        u1, u2, k1, k2 = p["u1"], p["u2"], p["k1"], p["k2"]
        if part_group == "ARKALIK" and any([u1, u2, k1, k2]):
            u1, u2, k1, k2 = False, False, False, False
            warnings.append("Arkalikta bant olamaz, kenar tikleri sifirlandi")

        db.add(
            OrderPart(
                id=str(uuid4()),
                order_id=order.id,
                part_group=part_group,
                boy_mm=p["boy_mm"],
                en_mm=p["en_mm"],
                adet=p["adet"],
                grain_code=p["grain_code"],
                u1=u1,
                u2=u2,
                k1=k1,
                k2=k2,
                part_desc=p["part_desc"],
                drill_code_1=p["drill_code_1"],
                drill_code_2=p["drill_code_2"],
            )
        )
        imported_rows += 1

    create_audit_log(db, current_user.id, "ORDER_CREATED_AUTO", f"Liste: {raw_name}", order.id)
    create_audit_log(
        db,
        current_user.id,
        "IMPORT_XLSX_AUTO",
        f"{imported_rows} satir import edildi ({part_group})",
        order.id,
    )
    db.commit()

    return {
        "order_id": str(order.id),
        "list_name": raw_name,
        "file_name": filename,
        "part_group": part_group,
        "imported_rows": imported_rows,
        "warnings": warnings,
        "status": "NEW",
    }


@router.post("/{order_id}/import/xlsx", status_code=200)
async def import_xlsx(
    order_id: str,
    file: UploadFile = File(...),
    part_group: str = Form("GOVDE"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_operator),
):
    """Excel dosyasından ölçü satırlarını siparişe ekle"""
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise NotFoundError("Sipariş")
    if order.status != "NEW":
        raise BusinessRuleError("Sadece NEW durumundaki siparişlere import yapılabilir")
    if part_group not in ("GOVDE", "ARKALIK"):
        raise BusinessRuleError("part_group GOVDE veya ARKALIK olmalı")

    # Dosya boyutu kontrolü (max 5MB)
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise ValidationError("Dosya boyutu 5MB'ı aşamaz")

    try:
        wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
        ws = wb.active

        imported_rows = 0
        warnings = []
        new_parts = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 3:
                continue

            try:
                boy_val = float(row[0]) if row[0] else 0
                en_val = float(row[1]) if row[1] else 0
                adet_val = int(row[2]) if row[2] else 1
            except (ValueError, TypeError):
                warnings.append(f"Satır {row_idx}: sayısal değer okunamadı, atlandı")
                continue

            if boy_val <= 0 or en_val <= 0:
                warnings.append(f"Satır {row_idx}: boy veya en sıfır/negatif, atlandı")
                continue

            if boy_val > 5000 or en_val > 5000:
                warnings.append(f"Satır {row_idx}: boy veya en 5000mm'den büyük, atlandı")
                continue

            if adet_val <= 0 or adet_val > 9999:
                warnings.append(f"Satır {row_idx}: adet geçersiz (1-9999 arası olmalı), atlandı")
                continue

            grain = str(row[3]).strip() if len(row) > 3 and row[3] else "0-Material"
            if grain not in ("0-Material", "1-Material", "2-Material", "3-Material"):
                grain = "0-Material"
                warnings.append(f"Satır {row_idx}: geçersiz grain, 0-Material kullanıldı")

            u1 = _parse_excel_bool(row[4]) if len(row) > 4 else False
            u2 = _parse_excel_bool(row[5]) if len(row) > 5 else False
            k1 = _parse_excel_bool(row[6]) if len(row) > 6 else False
            k2 = _parse_excel_bool(row[7]) if len(row) > 7 else False

            if part_group == "ARKALIK" and any([u1, u2, k1, k2]):
                warnings.append(f"Satır {row_idx}: arkalıkta bant olamaz, kenar tikleri sıfırlandı")
                u1, u2, k1, k2 = False, False, False, False
            desc = str(row[8]).strip() if len(row) > 8 and row[8] else None
            drill1 = str(row[9]).strip() if len(row) > 9 and row[9] else None
            drill2 = str(row[10]).strip() if len(row) > 10 and row[10] else None

            part = OrderPart(
                id=str(uuid4()),
                order_id=order_id,
                part_group=part_group,
                boy_mm=boy_val,
                en_mm=en_val,
                adet=adet_val,
                grain_code=grain,
                u1=u1,
                u2=u2,
                k1=k1,
                k2=k2,
                part_desc=desc,
                drill_code_1=drill1,
                drill_code_2=drill2,
            )
            db.add(part)
            new_parts.append(part)
            imported_rows += 1

        order.updated_at = datetime.now(timezone.utc)
        db.commit()

        create_audit_log(
            db,
            current_user.id,
            "IMPORT_XLSX",
            f"{imported_rows} satır import edildi ({part_group})",
            order_id,
        )
        db.commit()

        return {
            "imported_rows": imported_rows,
            "warnings": warnings,
            "parts": [OrderPartOut.model_validate(p) for p in new_parts],
        }

    except Exception as e:
        db.rollback()
        raise BusinessRuleError(f"Excel okuma hatası: {str(e)}")


# ═══════════════════════════════════════════════════
# SİPARİŞ SİLME - Sadece NEW durumunda
# ═══════════════════════════════════════════════════
@router.delete("/{order_id}", status_code=200)
def delete_order(
    order_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_operator),
):
    return OrderService.delete_order(db, order_id, current_user)


# ═══════════════════════════════════════════════════
@router.put("/{order_id}/parts", status_code=200)
def update_parts(
    order_id: str,
    parts: List[OrderPartCreate],
    db: Session = Depends(get_db),
    current_user: User = Depends(require_operator),
):
    order = OrderService.replace_parts(db, order_id, parts, current_user)
    return OrderService.order_to_out(order)


# ═══════════════════════════════════════════════════
@router.put("/{order_id}", response_model=OrderOut)
def update_order_header(
    order_id: str,
    body: OrderUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_operator),
):
    order = OrderService.update_order_header(db, order_id, body, current_user)
    return OrderService.order_to_out(order)
