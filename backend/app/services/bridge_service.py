"""
OptiPlan 360 — Bridge Service
Order + OrderParts → OptiPlanning XLSX (12-tag format)

Kurallar:
  - GOVDE ve ARKALIK ayrı dosyalar
  - ARKALIK'ta tüm kenar bantlar boş ("")
  - Birimler zaten mm (boy_mm, en_mm)
  - Grain: 0-Material→0, 1-Boyuna→1, 2-Enine→2, 3-Both→3
  - Band: u1=üst, u2=alt, k1=sol, k2=sağ kenar
  - Atomik yazma: .tmp → os.replace
"""

import io
import logging
import os
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)

# ── OptiPlanning 12 sütun tag'leri ──────────────────────────────────
OPTI_COLUMNS = [
    "[P_CODE_MAT]",  # Malzeme kodu (renk+kalınlık)
    "[P_LENGTH]",  # Uzunluk (mm)
    "[P_WIDTH]",  # Genişlik (mm)
    "[P_MINQ]",  # Adet
    "[P_GRAIN]",  # Desen yönü (0/1/2/3)
    "[P_IDESC]",  # Açıklama / trim bilgisi
    "[P_EDGE_UP]",  # Üst kenar bant (mm veya "")
    "[P_EDGE_LO]",  # Alt kenar bant (mm veya "")
    "[P_EDGE_SX]",  # Sol kenar bant (mm veya "")
    "[P_EDGE_DX]",  # Sağ kenar bant (mm veya "")
    "[P_CC_MAT]",  # Renk referansı
    "[P_DESC1]",  # Ek açıklama (part_desc)
]

GRAIN_MAP = {
    "0-Material": 0,
    "1-Boyuna": 1,
    "2-Enine": 2,
    "3-Both": 3,
}

EXPORT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)


def _material_code(order) -> str:
    """Malzeme kodu: renk + kalınlık. Örn: 'OAK18'"""
    color = (order.color or "UNKNOWN").replace(" ", "")
    thickness = int(float(order.thickness_mm)) if order.thickness_mm else 18
    return f"{color}{thickness}"


def _band_str(flag: bool, band_mm, is_arkalik: bool) -> str:
    """Kenar bant string değeri. ARKALIK'ta her zaman boş."""
    if is_arkalik or not flag or band_mm is None:
        return ""
    return str(float(band_mm))


def _trim_desc(order) -> str:
    """Trim açıklaması: 'XYmm gövde, Z.Zmm trim'"""
    thickness = float(order.thickness_mm) if order.thickness_mm else 18.0
    trim = round(thickness * 0.556, 1)
    return f"{int(thickness)}mm, {trim}mm trim"


def generate_optiplan_xlsx(order, parts: list) -> dict[str, bytes]:
    """
    Order + OrderPart listesinden OptiPlanning XLSX içeriği üretir.

    Returns:
        {"GOVDE": <bytes>, "ARKALIK": <bytes>} — sadece parça olan gruplar
    """
    result: dict[str, bytes] = {}
    mat_code = _material_code(order)
    cc_mat = order.color or ""
    trim_info = _trim_desc(order)

    for group_name in ("GOVDE", "ARKALIK"):
        is_arkalik = group_name == "ARKALIK"
        group_parts = [p for p in parts if (p.part_group or "").upper() == group_name]
        if not group_parts:
            continue

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = group_name

        # Başlık satırı
        ws.append(OPTI_COLUMNS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="FFD700")
            cell.alignment = Alignment(horizontal="center")

        for part in group_parts:
            grain = GRAIN_MAP.get(part.grain_code or "0-Material", 0)
            band_mm = order.band_mm
            adet = int(part.adet) if part.adet else 1

            ws.append(
                [
                    mat_code,
                    float(part.boy_mm or 0),
                    float(part.en_mm or 0),
                    adet,
                    grain,
                    trim_info,
                    _band_str(bool(part.u1), band_mm, is_arkalik),
                    _band_str(bool(part.u2), band_mm, is_arkalik),
                    _band_str(bool(part.k1), band_mm, is_arkalik),
                    _band_str(bool(part.k2), band_mm, is_arkalik),
                    cc_mat,
                    part.part_desc or "",
                ]
            )

        # Sütun genişliklerini otomatik ayarla
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        buf = io.BytesIO()
        wb.save(buf)
        result[group_name] = buf.getvalue()

    return result


def save_optiplan_xlsx(
    order,
    parts: list,
    output_dir: Optional[str] = None,
    job_id: Optional[str] = None,
) -> list[str]:
    """
    XLSX dosyalarını diske yazar (.tmp → rename atomik).

    Args:
        order: Order ORM nesnesi
        parts: OrderPart listesi
        output_dir: Hedef dizin (varsayılan: exports/)
        job_id: Dosya adına gömülen prefix (eşleştirme için)

    Returns:
        Yazılan dosya yollarının listesi
    """
    out_dir = output_dir or EXPORT_DIR
    os.makedirs(out_dir, exist_ok=True)

    xlsx_data = generate_optiplan_xlsx(order, parts)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    customer = (order.crm_name_snapshot or "unknown").replace(" ", "_")
    ts_code = (order.ts_code or "").replace(" ", "_")
    prefix = job_id[:8] if job_id else timestamp

    written: list[str] = []
    for group_name, content in xlsx_data.items():
        fname = f"{prefix}_{customer}_{ts_code}_{group_name}.xlsx"
        final_path = os.path.join(out_dir, fname)
        tmp_path = final_path + ".tmp"

        with open(tmp_path, "wb") as f:
            f.write(content)
        os.replace(tmp_path, final_path)

        part_count = sum(1 for p in parts if (p.part_group or "").upper() == group_name)
        logger.info("OptiPlanning XLSX yazıldı: %s (%d parça)", fname, part_count)
        written.append(final_path)

    return written


# ── Eski uyumluluk fonksiyonu (geriye dönük) ────────────────────────
def convert_excel_to_optiplanning(file_path: str) -> dict:
    """
    Excel dosyasını okuyup OptiPlanning veri yapısına çevirir.
    (Eski OCR import akışı için korunuyor)
    """
    try:
        import pandas as pd

        df = pd.read_excel(file_path)
        required = ["Boy", "En", "Adet", "Malzeme Adı", "Grup"]
        for col in required:
            if col not in df.columns:
                return {"error": f"Eksik sütun: {col}"}
        parts_list = [
            {
                "part_group": row["Grup"],
                "material_name": row["Malzeme Adı"],
                "thickness_mm": 18,
                "length_mm": row["Boy"],
                "width_mm": row["En"],
                "quantity": row["Adet"],
                "grain": "1-Boyuna",
                "description": row.get("Notlar", ""),
            }
            for _, row in df.iterrows()
        ]
        return {"customer_id": 1, "crm_name_snapshot": "Excel Import", "parts": parts_list}
    except Exception as exc:
        return {"error": str(exc)}
