import logging
import os
import shutil
import subprocess
from datetime import datetime, timezone
from typing import Any, Dict, List, Tuple

from app.exceptions import ValidationError as AppValidationError
from app.models import Order, OrderPart
from app.services.order_service import OrderService
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

# AGENT_ONEFILE_INSTRUCTIONS: Excel_sablon.xlsx tag'leri (sira ve isimlendirme LOCKED)
REQUIRED_SHEET_NAME = "ŞABLON"
REQUIRED_TAGS = [
    "[P_CODE_MAT]",
    "[P_LENGTH]",
    "[P_WIDTH]",
    "[P_MINQ]",
    "[P_GRAIN]",
    "[P_IDESC]",
    "[P_EDGE_MAT_UP]",
    "[P_EGDE_MAT_LO]",  # Yazim template'de boyle, duzeltilmez
    "[P_EDGE_MAT_SX]",
    "[P_EDGE_MAT_DX]",
    "[P_IIDESC]",
    "[P_DESC1]",
]

# AGENT_ONEFILE §THICKNESS POLICY: arkalik kalinliklari ve trim mapping
BACKING_THICKNESSES = frozenset([3, 4, 5, 8])
TRIM_BY_THICKNESS = {"18": 10.0, "3": 5.0, "4": 5.0, "5": 5.0, "8": 5.0}

# AGENT_ONEFILE §G4: bant mapping
EDGE_MAPPING = {
    "Bant Yok": None,
    "040": 0.4,
    "1mm": 1.0,
    "2mm": 2.0,
}


def _validate_template(template_path: str) -> Tuple[bool, str]:
    """Excel sablonunun gecerli oldugunu dogrular (AGENT_ONEFILE §3)."""
    if not os.path.exists(template_path):
        return False, f"Sablon dosyasi bulunamadi: {template_path}"
    try:
        import openpyxl

        wb = openpyxl.load_workbook(template_path, read_only=True)
        if REQUIRED_SHEET_NAME not in wb.sheetnames:
            return (
                False,
                f"Gerekli sheet bulunamadi: '{REQUIRED_SHEET_NAME}' (mevcut: {wb.sheetnames})",
            )
        ws = wb[REQUIRED_SHEET_NAME]
        first_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        first_row_trimmed = [str(v).strip() if v else "" for v in first_row[:12]]
        for i, expected_tag in enumerate(REQUIRED_TAGS):
            if i >= len(first_row_trimmed) or first_row_trimmed[i] != expected_tag:
                actual = first_row_trimmed[i] if i < len(first_row_trimmed) else "(eksik)"
                return (
                    False,
                    f"Tag uyumsuzlugu kolon {i+1}: beklenen='{expected_tag}', bulunan='{actual}'",
                )
        wb.close()
        return True, ""
    except Exception as e:
        return False, f"Sablon okunamadi: {e}"


def _map_edge_value(raw_value: Any) -> str:
    """Bant degerini OptiPlanning formatina donusturur (AGENT_ONEFILE §G4)."""
    if raw_value is None or raw_value is False or raw_value == "":
        return ""
    text = str(raw_value).strip()
    if text in EDGE_MAPPING:
        mapped = EDGE_MAPPING[text]
        return "" if mapped is None else str(mapped)
    # Boolean True -> bant var demek, ama spesifik deger yok
    if raw_value is True or text.lower() in ("true", "1", "evet"):
        return "1"
    return text


def _map_grain(raw_value: Any) -> int:
    """Grain mapping: 0/1/2/3 birebir (AGENT_ONEFILE §C3)."""
    if raw_value is None:
        return 0
    text = str(raw_value).strip()
    # "0-Material" gibi formattan sayi cikar
    if "-" in text:
        text = text.split("-")[0].strip()
    try:
        val = int(float(text))
        return val if val in (0, 1, 2, 3) else 0
    except (ValueError, TypeError):
        return 0


class OptiPlanningService:
    """
    OptiPlan 360 - Biesse OptiPlanning Integration Service
    Siparis verilerini AGENT_ONEFILE kurallarina gore Biesse formatina donusturur.
    XLSX uretimi SADECE Excel_sablon.xlsx template'i kopyalanarak yapilir (AGENT_ONEFILE §3).
    """

    def __init__(
        self,
        export_dir: str = os.environ.get("OPTIPLAN_EXPORT_DIR", r"C:\Biesse\OptiPlanning\Tmp\Sol"),
        optiplan_exe: str = os.environ.get(
            "OPTIPLAN_EXE_PATH", r"C:\Biesse\OptiPlanning\System\OptiPlanning.exe"
        ),
    ):
        self.export_dir = self._resolve_writable_export_dir(export_dir)
        self.optiplan_exe = optiplan_exe
        self.template_path = self._resolve_template_path()

    def _resolve_writable_export_dir(self, preferred_dir: str) -> str:
        """Yazma izni olan export klasorunu sec; gerekirse proje ici fallback kullan."""
        project_tmp = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))),
            "tmp",
            "optiplan_exports",
        )
        candidates = [preferred_dir, project_tmp]

        for candidate in candidates:
            try:
                os.makedirs(candidate, exist_ok=True)
                probe = os.path.join(candidate, ".write_test")
                with open(probe, "w", encoding="utf-8") as f:
                    f.write("ok")
                os.remove(probe)
                return candidate
            except Exception:
                continue

        raise AppValidationError(
            "E_EXPORT_DIR_UNWRITABLE: Export klasoru yazilabilir degil. "
            f"Denenen yollar: {candidates}"
        )

    # AGENT_ONEFILE_INSTRUCTIONS §3: template repo konumu    def _resolve_template_path(self) -> str:
    def _resolve_template_path(self) -> str:
        """
        Excel template yolunu runtime'da cozer.
        Sirayla:
        1) OPTIPLAN_TEMPLATE_PATH (varsa),
        2) backend/templates/Excel_sablon.xlsx,
        3) repo_root/templates/Excel_sablon.xlsx,
        4) cwd/templates/Excel_sablon.xlsx.
        """
        env_path = os.environ.get("OPTIPLAN_TEMPLATE_PATH", "").strip()
        backend_root = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
        repo_root = os.path.dirname(backend_root)

        candidates: list[str] = []
        if env_path:
            candidates.append(env_path)

        candidates.extend(
            [
                os.path.join(backend_root, "templates", "Excel_sablon.xlsx"),
                os.path.join(repo_root, "templates", "Excel_sablon.xlsx"),
                os.path.join(os.getcwd(), "templates", "Excel_sablon.xlsx"),
            ]
        )

        for candidate in candidates:
            if os.path.exists(candidate):
                return candidate

        # Hata mesaji net olsun diye once env path'i, yoksa ilk fallback'i dondur.
        return env_path or candidates[0]

    def export_order(
        self, db: Session, order_id: str, trigger_exe: bool = False, format_type: str = "EXCEL"
    ) -> List[str]:
        """
        Ana disa aktarma (export) fonksiyonu. Siparisi alip GOVDE/ARKALIK ve
        renk/kalinlik bazinda ayirip dosyalar olusturur.
        """
        # Template validasyonu (AGENT_ONEFILE §3)
        if format_type.upper() == "EXCEL":
            valid, err = _validate_template(self.template_path)
            if not valid:
                raise AppValidationError(f"E_TEMPLATE_INVALID: {err}")

        order = OrderService.get_order(db, order_id, with_parts=True)

        if not order.parts:
            raise ValueError(f"Siparis ({order_id}) icin disa aktarilacak parca bulunamadi.")

        generated_files = []

        # 1. Parcalari gruplama (Govde vs Arkalik) ve (Kalinlik + Renk)
        grouped_parts = self._group_parts(order)

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        crm_name = order.crm_name_snapshot or f"CUST_{order.customer_id}"

        # CRMISIM_TIMESTAMP_18mmBeyaz_GOVDE.xlsx
        safe_crm_name = "".join(c for c in crm_name if c.isalnum() or c in (" ", "_", "-")).replace(
            " ", ""
        )

        for group_key, parts in grouped_parts.items():
            meta = dict(group_key)
            thickness = meta["thickness"]
            color = meta["color"]
            part_group = meta["part_group"]

            safe_color = "".join(c for c in color if c.isalnum() or c in (" ", "_", "-")).replace(
                " ", ""
            )

            # Kalinlik degerini integer yap (18.0 -> 18) — dosya adinda nokta OptiPlanning'de sorun yaratir
            thick_int = int(float(thickness)) if thickness else 18
            filename = f"{safe_crm_name}_{timestamp}_{thick_int}mm{safe_color}_{part_group}"

            if format_type.upper() == "EXCEL":
                filepath = self._generate_excel(parts, filename, part_group, int(thickness))
                if filepath:
                    generated_files.append(filepath)
            elif format_type.upper() == "OSI":
                logger.warning("OSI export is pending implementation")
            elif format_type.upper() == "XML":
                logger.warning("XML export is pending implementation")

        if trigger_exe and generated_files and os.path.exists(self.optiplan_exe):
            self._trigger_optiplan(generated_files)

        return generated_files

    def _group_parts(self, order: Order) -> Dict[Any, List[OrderPart]]:
        """
        Parcalari Renk, Kalinlik ve Parca Grubuna (GOVDE/ARKALIK) gore gruplar.
        AGENT_ONEFILE §THICKNESS POLICY: arkalik kalinliklari [3, 4, 5, 8]
        """
        groups: Dict[Any, List[OrderPart]] = {}
        for part in order.parts:
            part_group_raw = getattr(part, "part_group", "GOVDE") or "GOVDE"
            part_group = part_group_raw.upper()

            if part_group == "GOVDE":
                thickness = float(order.thickness_mm) if order.thickness_mm else 18
            else:
                # Arkalik: varsayilan 8, ama part'ta ozel kalinlik varsa onu kullan
                thickness = 8

            color = order.color or "Standart"

            key = frozenset(
                {"thickness": thickness, "color": color, "part_group": part_group}.items()
            )

            if key not in groups:
                groups[key] = []
            groups[key].append(part)

        return groups

    def _generate_excel(
        self, parts: List[OrderPart], filename_prefix: str, part_group: str, thickness: int = 18
    ) -> str:
        """
        AGENT_ONEFILE §3: XLSX uretimi SADECE Excel_sablon.xlsx template'i
        kopyalanarak yapilir. Tag satiri (row 1) ve baslik satiri (row 2)
        degistirilmez, veriler 3. satirdan itibaren yazilir.
        """
        import openpyxl

        is_arkalik = part_group.upper() == "ARKALIK"
        thickness_key = str(int(thickness))
        trim_value = TRIM_BY_THICKNESS.get(thickness_key, 10.0)

        # openpyxl uzanti kontrolu yaptigi icin gecici dosya da .xlsx olmalidir.
        # Atomik yazma: .tmp.xlsx -> .xlsx
        tmp_filepath = os.path.join(self.export_dir, f"{filename_prefix}.tmp.xlsx")
        final_filepath = os.path.join(self.export_dir, f"{filename_prefix}.xlsx")

        try:
            shutil.copy2(self.template_path, tmp_filepath)

            wb = openpyxl.load_workbook(tmp_filepath)
            ws = wb[REQUIRED_SHEET_NAME]

            # Sablondaki olasi ornek veri satirlarini temizle (row 1-2 korunur).
            if ws.max_row >= 3:
                for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=12):
                    for cell in row:
                        cell.value = None

            # Veriler 3. satirdan itibaren (row 1=tags, row 2=basliklar)
            for row_idx, part in enumerate(parts, start=3):
                # [P_CODE_MAT] - Malzeme kodu
                ws.cell(
                    row=row_idx,
                    column=1,
                    value=getattr(part, "id", "")[-6:] if getattr(part, "id", "") else "",
                )

                # [P_LENGTH] - Boy (mm)
                ws.cell(row=row_idx, column=2, value=float(part.boy_mm) if part.boy_mm else 0)

                # [P_WIDTH] - En (mm)
                ws.cell(row=row_idx, column=3, value=float(part.en_mm) if part.en_mm else 0)

                # [P_MINQ] - Adet
                ws.cell(row=row_idx, column=4, value=int(part.adet) if part.adet else 1)

                # [P_GRAIN] - Grain mapping 0/1/2/3 (AGENT_ONEFILE §C3)
                grain_val = getattr(part, "grain_code", None) or getattr(part, "grain", "0")
                ws.cell(row=row_idx, column=5, value=_map_grain(grain_val))

                # [P_IDESC] - Trim degeri (AGENT_ONEFILE §THICKNESS POLICY)
                ws.cell(row=row_idx, column=6, value=trim_value)

                # AGENT_ONEFILE §C4: Arkalikta bant her kosulda null
                if is_arkalik:
                    # [P_EDGE_MAT_UP], [P_EGDE_MAT_LO], [P_EDGE_MAT_SX], [P_EDGE_MAT_DX]
                    ws.cell(row=row_idx, column=7, value="")
                    ws.cell(row=row_idx, column=8, value="")
                    ws.cell(row=row_idx, column=9, value="")
                    ws.cell(row=row_idx, column=10, value="")
                else:
                    ws.cell(row=row_idx, column=7, value=_map_edge_value(part.u1))
                    ws.cell(row=row_idx, column=8, value=_map_edge_value(part.u2))
                    ws.cell(row=row_idx, column=9, value=_map_edge_value(part.k1))
                    ws.cell(row=row_idx, column=10, value=_map_edge_value(part.k2))

                # [P_IIDESC] - Drill code / ek kod
                ws.cell(row=row_idx, column=11, value=getattr(part, "drill_code_1", "") or "")

                # [P_DESC1] - Aciklama
                ws.cell(row=row_idx, column=12, value=part.part_desc or "")

            wb.save(tmp_filepath)
            wb.close()

            # Atomik replace: .tmp.xlsx -> .xlsx
            os.replace(tmp_filepath, final_filepath)
            logger.info("Excel basariyla olusturuldu (template-based): %s", final_filepath)
            return final_filepath

        except Exception as e:
            logger.error("Excel dosyasi olusturulurken hata: %s", e)
            if os.path.exists(tmp_filepath):
                os.remove(tmp_filepath)
            raise

    def _trigger_optiplan(self, file_paths: List[str]) -> None:
        """OptiPlan.exe uygulamasini belirtilen dosyalar icin tetikler."""
        try:
            for path in file_paths:
                subprocess.Popen([self.optiplan_exe, path], shell=False)
            logger.info("OptiPlan exe tetiklendi: %d dosya.", len(file_paths))
        except Exception as e:
            logger.error("OptiPlan tetiklenirken hata: %s", e)

    # --- Advanced Features: Machine Config ---
    def get_machine_config(self, db: Session, config_name: str = "DEFAULT"):
        """Belirtilen ada sahip makine konfigurasyonunu getirir."""
        from app.models.optiplanning import MachineConfig

        return db.query(MachineConfig).filter(MachineConfig.name == config_name).first()

    def update_machine_config(self, db: Session, config_data: dict, config_name: str = "DEFAULT"):
        """Makine konfigurasyonunu gunceller veya yoksa olusturur."""
        from app.models.optiplanning import MachineConfig

        config = self.get_machine_config(db, config_name)

        if not config:
            config = MachineConfig(name=config_name, **config_data)
            db.add(config)
        else:
            for key, value in config_data.items():
                if hasattr(config, key) and value is not None:
                    setattr(config, key, value)

        db.commit()
        db.refresh(config)
        return config

    def run_advanced_optimization(self, db: Session, job_id: str, request_data: Any):
        """Gelismis optimizasyon isini calistirir."""
        from app.models.optiplanning import OptimizationJob

        job = db.query(OptimizationJob).filter(OptimizationJob.id == job_id).first()
        if not job:
            raise ValueError("Optimizasyon isi bulunamadi.")

        try:
            job.status = "RUNNING"
            db.commit()

            generated_files = []
            for order_id in request_data.order_ids:
                files = self.export_order(
                    db=db, order_id=str(order_id), trigger_exe=False, format_type="EXCEL"
                )
                generated_files.extend(files)

            if generated_files and request_data.params:
                self._trigger_optiplan(generated_files)

            job.status = "COMPLETED"
            job.result_file_path = ",".join(generated_files) if generated_files else None
            job.completed_at = datetime.now(timezone.utc)
            db.commit()
            db.refresh(job)

            return job
        except Exception as e:
            job.status = "FAILED"
            job.error_message = str(e)
            job.completed_at = datetime.now(timezone.utc)
            db.commit()
            raise e


optiplanning_service = OptiPlanningService()
