"""
Fiyat Takip Servisi — dosya yükleme, işleme, birleştirme, export.
İş mantığı bu katmanda; router sadece HTTP in/out yapar.
"""
import io
import logging
import tempfile
from datetime import datetime
from pathlib import Path
from uuid import uuid4

import pandas as pd
from fastapi import BackgroundTasks
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy.orm import Session, sessionmaker

from app.exceptions import (
    AuthorizationError,
    BusinessRuleError,
    NotFoundError,
    ValidationError,
)
from app.models import PriceItem, PriceJobStatusEnum, PriceUploadJob, User
from app.services.price_tracking_ai import extract_price_data_from_text
from app.services.price_tracking_helpers import (
    MAX_FILE_SIZE_MB,
    SUPPORTED_EXTENSIONS,
    get_file_type,
    load_mapping_rules,
    normalize_columns,
)
from app.services.price_tracking_ocr import (
    extract_text_from_image,
    extract_text_from_pdf_images,
)

logger = logging.getLogger(__name__)


class PriceTrackingService:
    """Fiyat listesi işleme servisi."""

    # ── Yetki kontrolü ─────────────────────────────────────

    @staticmethod
    def _assert_can_modify(job: PriceUploadJob, user: User) -> None:
        role = (user.role or "").upper()
        if role == "ADMIN":
            return
        if role in ("OPERATOR", "SALES"):
            if job.uploaded_by_id != user.id:
                raise AuthorizationError(
                    "Yalnızca kendi yüklediğiniz işleri değiştirebilirsiniz"
                )
            return
        raise AuthorizationError("Bu işlem için yetersiz rol")

    # ── CRUD ───────────────────────────────────────────────

    @staticmethod
    def list_jobs(db: Session, user: User) -> list[PriceUploadJob]:
        """Kullanıcının erişebildiği tüm job'ları listeler."""
        role = (user.role or "").upper()
        query = db.query(PriceUploadJob).order_by(
            PriceUploadJob.created_at.desc()
        )
        if role != "ADMIN":
            query = query.filter(PriceUploadJob.uploaded_by_id == user.id)
        return query.all()

    @staticmethod
    def get_job(db: Session, job_id: str) -> PriceUploadJob:
        """Job'ı getirir; bulamazsa NotFoundError."""
        job = db.query(PriceUploadJob).filter(PriceUploadJob.id == job_id).first()
        if not job:
            raise NotFoundError("Fiyat yükleme işi", job_id)
        return job

    @staticmethod
    def get_job_detail(db: Session, job_id: str, user: User) -> PriceUploadJob:
        """Job + items detay."""
        job = PriceTrackingService.get_job(db, job_id)
        role = (user.role or "").upper()
        if role != "ADMIN" and job.uploaded_by_id != user.id:
            raise AuthorizationError("Bu işe erişim yetkiniz yok")
        return job

    @staticmethod
    def delete_job(db: Session, job_id: str, user: User) -> None:
        """Job ve ilişkili item'ları siler."""
        job = PriceTrackingService.get_job(db, job_id)
        PriceTrackingService._assert_can_modify(job, user)
        db.delete(job)
        db.commit()
        logger.info("Fiyat işi silindi: %s (kullanıcı: %s)", job_id, user.id)

    # ── Dosya yükleme + işleme ─────────────────────────────

    @staticmethod
    def upload_and_process(
        db: Session,
        file_data: bytes,
        filename: str,
        content_type: str,
        supplier: str,
        user: User,
        background_tasks: BackgroundTasks | None = None,
    ) -> PriceUploadJob:
        """
        Dosya yukler, job olusturur ve isleme surecini baslatir.

        Raises:
            ValidationError: Geçersiz dosya tipi/boyutu
        """
        ext = Path(filename).suffix.lower()
        if ext not in SUPPORTED_EXTENSIONS:
            raise ValidationError(
                f"Desteklenmeyen dosya formatı: {ext}. "
                f"Desteklenen: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
            )

        file_size_mb = len(file_data) / (1024 * 1024)
        if file_size_mb > MAX_FILE_SIZE_MB:
            raise ValidationError(
                f"Dosya boyutu limiti aşıldı: {file_size_mb:.1f}MB > {MAX_FILE_SIZE_MB}MB"
            )

        # Job oluştur
        job = PriceUploadJob(
            id=str(uuid4()),
            status=PriceJobStatusEnum.PENDING.value,
            original_filename=filename,
            content_type=content_type,
            file_size=len(file_data),
            file_data=file_data,
            supplier=supplier,
            uploaded_by_id=user.id,
        )
        db.add(job)
        db.commit()
        db.refresh(job)

        # Dependency override edilen test ortamlarinda da ayni engine kullanilsin.
        session_factory = sessionmaker(
            bind=db.get_bind(),
            autocommit=False,
            autoflush=False,
        )
        if background_tasks is not None:
            background_tasks.add_task(
                PriceTrackingService._process_file,
                job.id,
                session_factory,
            )
        else:
            PriceTrackingService._process_file(job.id, session_factory)

        return job

    @staticmethod
    def _process_file(job_id: str, session_factory: sessionmaker) -> None:
        """Job ID uzerinden dosyayi isler (background-task uyumlu)."""
        db = session_factory()
        try:
            job = PriceTrackingService.get_job(db, job_id)
            job.status = PriceJobStatusEnum.PROCESSING.value
            job.error_message = None
            db.commit()
            file_data = job.file_data or b""
            PriceTrackingService._process_job(db, job, file_data)
        except Exception as e:
            db.rollback()
            try:
                job = db.query(PriceUploadJob).filter(PriceUploadJob.id == job_id).first()
                if job:
                    job.status = PriceJobStatusEnum.FAILED.value
                    job.error_message = str(e)
                    db.commit()
            except Exception:
                db.rollback()
            logger.error("Fiyat isleme hatasi (job=%s): %s", job_id, e)
        finally:
            db.close()

    @staticmethod
    def _process_job(db: Session, job: PriceUploadJob, file_data: bytes) -> None:
        """İş dosyasını tipine göre işler ve item'ları DB'ye yazar."""
        ext = Path(job.original_filename).suffix.lower()
        file_type = get_file_type(ext)

        # Geçici dosyaya yaz (kütüphaneler dosya yolu istiyor)
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            tmp.write(file_data)
            tmp_path = tmp.name

        try:
            if file_type == "spreadsheet":
                df = PriceTrackingService._process_spreadsheet(tmp_path, job.supplier)
            elif file_type == "pdf":
                df = PriceTrackingService._process_pdf(tmp_path, job.supplier)
            elif file_type == "image":
                df = PriceTrackingService._process_image(tmp_path, job.supplier)
            else:
                raise BusinessRuleError(f"Desteklenmeyen dosya tipi: {file_type}")

            if df.empty:
                job.status = PriceJobStatusEnum.COMPLETED.value
                job.rows_extracted = 0
                job.error_message = "Dosyadan veri çıkarılamadı"
                db.commit()
                return

            # Hesaplanabilir alanları doldur
            df = PriceTrackingService._calculate_derived_fields(df)

            # DB'ye yaz
            items_created = 0
            for _, row in df.iterrows():
                item = PriceItem(
                    id=str(uuid4()),
                    job_id=job.id,
                    urun_kodu=_safe_str(row.get("URUN_KODU")),
                    urun_adi=str(row.get("URUN_ADI", "")),
                    birim=str(row.get("BIRIM", "ADET")),
                    boyut=_safe_str(row.get("BOYUT")),
                    renk=_safe_str(row.get("RENK")),
                    liste_fiyati=_safe_float(row.get("LISTE_FIYATI")),
                    iskonto_orani=_safe_float(row.get("ISKONTO_ORANI"), 0),
                    net_fiyat=_safe_float(row.get("NET_FIYAT")),
                    kdv_orani=_safe_float(row.get("KDV_ORANI"), 20),
                    kdv_dahil_fiyat=_safe_float(row.get("KDV_DAHIL_FIYAT")),
                    para_birimi=str(row.get("PARA_BIRIMI", "TRY")),
                    kategori=_safe_str(row.get("KATEGORI")),
                    marka=_safe_str(row.get("MARKA")),
                    tedarikci=job.supplier,
                )
                db.add(item)
                items_created += 1

            job.status = PriceJobStatusEnum.COMPLETED.value
            job.rows_extracted = items_created
            db.commit()
            logger.info(
                "Fiyat işleme tamamlandı: job=%s, %d ürün", job.id, items_created
            )

        finally:
            Path(tmp_path).unlink(missing_ok=True)

    # ── Format-spesifik işleme ─────────────────────────────

    @staticmethod
    def _process_spreadsheet(file_path: str, supplier: str) -> pd.DataFrame:
        """Excel dosyasını okur ve standart sütunlara eşler."""
        try:
            df = pd.read_excel(file_path, engine="openpyxl")
        except Exception:
            df = pd.read_excel(file_path)

        if df.empty:
            return pd.DataFrame()

        # DEBUG: Orijinal sütunları logla
        logger.info("EXCEL ORIJINAL SUTUNLAR: %s", list(df.columns))
        logger.info("EXCEL SUTUN UNIQUE MI: %s", df.columns.is_unique)

        col_mapping = normalize_columns(df.columns.tolist())
        logger.info("ESLEME SONUCU: %s", col_mapping)
        
        if col_mapping:
            df = df.rename(columns=col_mapping)
            
            # DEBUG: Rename sonrası sütunları logla
            logger.info("RENAME SONRASI SUTUNLAR: %s", list(df.columns))
            logger.info("RENAME SONRASI UNIQUE MI: %s", df.columns.is_unique)
            
            # Duplicate sütunları temizle (güvenlik katmanı)
            if not df.columns.is_unique:
                logger.warning("DUPLICATE SUTUN TESPIT EDILDI! Temizleniyor...")
                # Her sütun grubu için son geçerli değeri al
                df = df.T.groupby(level=0).last().T
                logger.info("TEMIZLIK SONRASI SUTUNLAR: %s", list(df.columns))
            
            # Sayisal donusumleri burada da yap
            df = PriceTrackingService._normalize_data_types(df)
            
            # ── Renk ayrıştırma: "/" ile ayrılmış çoklu renkleri ayrı satırlara aç ──
            df = PriceTrackingService._explode_multi_color(df)
            
            # DEBUG: İlk satırı logla
            if len(df) > 0:
                first_row = df.iloc[0]
                for col in df.columns:
                    val = first_row[col]
                    logger.info("  SUTUN[%s] = %r (type=%s)", col, val, type(val).__name__)
        else:
            # Sütun eşleme yapılamazsa AI'ya gönder
            text = df.to_string(index=False)
            items = extract_price_data_from_text(text, supplier=supplier)
            if items:
                df = pd.DataFrame(items)
                df = PriceTrackingService._normalize_ai_output(df)

        return df

    @staticmethod
    def _process_pdf(file_path: str, supplier: str) -> pd.DataFrame:
        """PDF'den tablo veya metin çıkarır, AI ile yapılandırır."""
        try:
            import pdfplumber
        except ImportError:
            logger.error("pdfplumber yüklü değil")
            return pd.DataFrame()

        # pdfplumber ile tablo çıkarımı dene
        all_rows: list = []
        headers: list[str] | None = None

        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    for table in page.extract_tables():
                        if not table:
                            continue
                        for row_idx, row in enumerate(table):
                            if row_idx == 0 and headers is None:
                                headers = [str(c or "").strip() for c in row]
                            else:
                                all_rows.append(row)
        except Exception as e:
            logger.warning("pdfplumber tablo çıkarımı başarısız: %s", e)

        if all_rows and headers:
            df = pd.DataFrame(all_rows, columns=headers).dropna(how="all")
            col_mapping = normalize_columns(df.columns.tolist())
            if col_mapping:
                df = df.rename(columns=col_mapping)
                return df

        # Metin çıkar → AI
        text_parts: list[str] = []
        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text_parts.append(page_text)
        except Exception:
            pass

        if not text_parts:
            text = extract_text_from_pdf_images(file_path)
        else:
            text = "\n".join(text_parts)

        items = extract_price_data_from_text(text, supplier=supplier)
        if items:
            df = pd.DataFrame(items)
            return PriceTrackingService._normalize_ai_output(df)

        return pd.DataFrame()

    @staticmethod
    def _process_image(file_path: str, supplier: str) -> pd.DataFrame:
        """Görüntüden OCR + AI ile fiyat verisi çıkarır."""
        text = extract_text_from_image(file_path)

        if not text or len(text.strip()) < 20:
            return pd.DataFrame()

        items = extract_price_data_from_text(text, supplier=supplier)
        if items:
            df = pd.DataFrame(items)
            return PriceTrackingService._normalize_ai_output(df)

        return pd.DataFrame()

    # ── AI çıktı normalizasyonu ────────────────────────────

    @staticmethod
    def _normalize_data_types(df: pd.DataFrame) -> pd.DataFrame:
        """Veri tiplerini (sayısal, string) standartlaştırır."""
        # Sayısal dönüşüm
        for col in ("LISTE_FIYATI", "ISKONTO_ORANI", "NET_FIYAT", "KDV_ORANI", "KDV_DAHIL_FIYAT"):
            if col in df.columns:
                # Önce para birimi sembollerini temizle (örn: "1.250,00 TL" -> 1250.00)
                # Basit bir temizlik: sadece rakam, nokta, virgül ve eksi kalsın
                if df[col].dtype == "object":
                   # Virgül ondalık ise noktaya çevir, binlik ayırıcı noktaları kaldır (Tr stili)
                   # Ancak format belirsiz olabilir. pd.to_numeric "coerce" en güvenlisi.
                   # Karmaşık string temizliği gerekebilir, şimdilik basit coerce:
                   pass
                
                df[col] = pd.to_numeric(df[col], errors="coerce")
        
        # String temizliği - Sadece numeric olmayan kolonlar
        for col in df.columns:
            if col not in ("LISTE_FIYATI", "ISKONTO_ORANI", "NET_FIYAT", "KDV_ORANI", "KDV_DAHIL_FIYAT"):
                if df[col].dtype == "object":
                    df[col] = df[col].astype(str).str.strip()

        return df

    @staticmethod
    def _normalize_ai_output(df: pd.DataFrame) -> pd.DataFrame:
        """AI'dan gelen sütun isimlerini standart formata dönüştürür."""
        rename_map = {
            "urun_kodu": "URUN_KODU",
            "urun_adi": "URUN_ADI",
            "birim": "BIRIM",
            "liste_fiyati": "LISTE_FIYATI",
            "iskonto_orani": "ISKONTO_ORANI",
            "net_fiyat": "NET_FIYAT",
            "kdv_orani": "KDV_ORANI",
            "kdv_dahil_fiyat": "KDV_DAHIL_FIYAT",
            "para_birimi": "PARA_BIRIMI",
            "stok_durumu": "STOK_DURUMU",
            "kategori": "KATEGORI",
            "marka": "MARKA",
            "aciklama": "ACIKLAMA",
        }
        existing = {k: v for k, v in rename_map.items() if k in df.columns}
        df = df.rename(columns=existing)

        # Veri tipi dönüşümü
        df = PriceTrackingService._normalize_data_types(df)

        # Birim normalizasyonu
        rules = load_mapping_rules()
        unit_map = rules.get("unit_normalization", {})
        if "BIRIM" in df.columns:
            df["BIRIM"] = df["BIRIM"].map(
                lambda x: unit_map.get(str(x).strip(), str(x).strip()) if pd.notna(x) else "ADET"
            )

        # Para birimi normalizasyonu
        currency_map = rules.get("currency_symbols", {})
        if "PARA_BIRIMI" in df.columns:
            df["PARA_BIRIMI"] = df["PARA_BIRIMI"].map(
                lambda x: currency_map.get(str(x).strip(), str(x).strip()) if pd.notna(x) else "TRY"
            )

        return df

    # ── Hesaplanabilir alanlar ─────────────────────────────

    @staticmethod
    def _explode_multi_color(df: pd.DataFrame) -> pd.DataFrame:
        """
        Çoklu renk ayrıştırma kuralı:
        
        RENK sütununda '/' ile ayrılmış değerler varsa (ör: 'ALTIN/BRONZ/FIR.ANTİK/MAT ALTIN/MAT FÜME'),
        her renk için ayrı bir satır oluşturur. Her satır orijinal ürünün tüm verilerini taşır.
        
        İş kuralı: 32'nin katları olan sayılar (32, 64, 96, 128, 160, 192, 224, 256, 288, 320, 352...)
        ürün boyutudur (mm cinsinden).
        """
        if "RENK" not in df.columns:
            return df
        
        new_rows = []
        exploded = False
        
        for _, row in df.iterrows():
            renk_val = row.get("RENK")
            
            # Renk değeri yoksa veya NaN ise olduğu gibi bırak
            if pd.isna(renk_val) or not str(renk_val).strip():
                new_rows.append(row.to_dict())
                continue
            
            renk_str = str(renk_val).strip()
            
            # '/' ile ayrılmış çoklu renkler var mı?
            if "/" in renk_str:
                colors = [c.strip() for c in renk_str.split("/") if c.strip()]
                
                if len(colors) > 1:
                    exploded = True
                    for color in colors:
                        new_row = row.to_dict()
                        new_row["RENK"] = color
                        new_rows.append(new_row)
                else:
                    new_rows.append(row.to_dict())
            else:
                new_rows.append(row.to_dict())
        
        if exploded:
            result = pd.DataFrame(new_rows)
            logger.info(
                "Renk ayrıştırma: %d satır → %d satır (çoklu renkler açıldı)",
                len(df), len(result)
            )
            return result.reset_index(drop=True)
        
        return df

    @staticmethod
    def _calculate_derived_fields(df: pd.DataFrame) -> pd.DataFrame:
        """NET_FIYAT ve KDV_DAHIL_FIYAT hesaplar."""
        
        # ── Güvenlik: Duplicate sütunları temizle ──
        if not df.columns.is_unique:
            df = df.T.groupby(level=0).last().T
        
        # ── Güvenlik: Sayısal sütunları zorla sayısal yap ──
        for col in ("LISTE_FIYATI", "ISKONTO_ORANI", "NET_FIYAT", "KDV_ORANI", "KDV_DAHIL_FIYAT"):
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        
        # Default değerler
        if "ISKONTO_ORANI" not in df.columns:
            df["ISKONTO_ORANI"] = 0
        if "KDV_ORANI" not in df.columns:
            df["KDV_ORANI"] = 20

        df["ISKONTO_ORANI"] = df["ISKONTO_ORANI"].fillna(0)
        df["KDV_ORANI"] = df["KDV_ORANI"].fillna(20)

        # Net fiyat
        if "LISTE_FIYATI" in df.columns:
            mask = (
                df["NET_FIYAT"].isna()
                if "NET_FIYAT" in df.columns
                else pd.Series(True, index=df.index)
            )
            df.loc[mask, "NET_FIYAT"] = (
                df.loc[mask, "LISTE_FIYATI"] * (1 - df.loc[mask, "ISKONTO_ORANI"] / 100)
            )

        # KDV dahil fiyat
        if "NET_FIYAT" in df.columns:
            mask = (
                df["KDV_DAHIL_FIYAT"].isna()
                if "KDV_DAHIL_FIYAT" in df.columns
                else pd.Series(True, index=df.index)
            )
            df.loc[mask, "KDV_DAHIL_FIYAT"] = (
                df.loc[mask, "NET_FIYAT"] * (1 + df.loc[mask, "KDV_ORANI"] / 100)
            )

        # Ürün adı boş satırları kaldır
        if "URUN_ADI" in df.columns:
            df = df.dropna(subset=["URUN_ADI"])
            df = df[df["URUN_ADI"].astype(str).str.strip() != ""]

        return df

    # ── Excel Export ───────────────────────────────────────

    @staticmethod
    def export_to_excel(
        db: Session, job_ids: list[str], user: User
    ) -> bytes:
        """Seçili job'ların ürünlerini birleştirilmiş Excel dosyasına çevirir."""
        all_items: list[PriceItem] = []

        for job_id in job_ids:
            job = PriceTrackingService.get_job(db, job_id)
            role = (user.role or "").upper()
            if role != "ADMIN" and job.uploaded_by_id != user.id:
                raise AuthorizationError(f"İş {job_id} için erişim yetkiniz yok")
            all_items.extend(job.items)

        if not all_items:
            raise BusinessRuleError("Seçili işlerde dışa aktarılacak ürün yok")

        # DataFrame oluştur
        data = []
        for item in all_items:
            data.append({
                "URUN_KODU": item.urun_kodu or "",
                "URUN_ADI": item.urun_adi,
                "BIRIM": item.birim,
                "LISTE_FIYATI": float(item.liste_fiyati) if item.liste_fiyati else None,
                "ISKONTO_ORANI": float(item.iskonto_orani) if item.iskonto_orani else 0,
                "NET_FIYAT": float(item.net_fiyat) if item.net_fiyat else None,
                "KDV_ORANI": float(item.kdv_orani) if item.kdv_orani else 20,
                "KDV_DAHIL_FIYAT": float(item.kdv_dahil_fiyat) if item.kdv_dahil_fiyat else None,
                "PARA_BIRIMI": item.para_birimi,
                "KATEGORI": item.kategori or "",
                "MARKA": item.marka or "",
                "TEDARIKCI": item.tedarikci,
            })

        df = pd.DataFrame(data)

        # Formatlı Excel oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Fiyat Listesi"

        # Stiller
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

        # Header
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

        # Kolon genişlikleri
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(len(str(col_name)), 10)
            letter = chr(64 + col_idx) if col_idx <= 26 else "A"
            ws.column_dimensions[letter].width = min(max_len + 4, 35)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        # Özet sayfası
        _add_summary_sheet(wb, df)

        # Bytes olarak döndür
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    @staticmethod
    def export_jobs_to_excel(
        db: Session, job_ids: list[str], user: User
    ) -> bytes:
        """Geriye uyumluluk için eski export metodu."""
        return PriceTrackingService.export_to_excel(db, job_ids, user)


# ── Yardımcı fonksiyonlar ─────────────────────────────────

def _extract_scalar(value):
    """
    Pandas Series/DataFrame gibi çoklu değer dönebilen yapıların içinden 
    tekil (skalar) değeri çıkarır. Defensive programming.
    """
    if isinstance(value, (pd.Series, pd.DataFrame)):
        if value.empty:
            return None
        # İlk geçerli değeri almaya çalış, yoksa ilk değeri al
        valid = value.dropna()
        if not valid.empty:
            return valid.iloc[0]
        return value.iloc[0]
    return value


def _safe_str(value) -> str | None:
    value = _extract_scalar(value)
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    return str(value).strip() or None


def _safe_float(value, default=None) -> float | None:
    value = _extract_scalar(value)
    if value is None:
        return default
    try:
        f = float(value)
        return f if not pd.isna(f) else default
    except (ValueError, TypeError):
        return default


def _add_summary_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """Özet istatistik sayfası ekler."""
    ws = wb.create_sheet("Özet")
    
    # Duplicate sütun kontrolü (defensive)
    if not df.columns.is_unique:
        df = df.loc[:, ~df.columns.duplicated(keep='last')]

    summary = [
        ("Toplam Ürün Sayısı", len(df)),
        ("Tedarikçi Sayısı", df["TEDARIKCI"].nunique() if "TEDARIKCI" in df.columns else 0),
        ("İşlem Tarihi", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]

    if "LISTE_FIYATI" in df.columns:
        valid = df["LISTE_FIYATI"].dropna()
        if not valid.empty:
            summary.extend([
                ("", ""),
                ("Fiyat İstatistikleri", ""),
                ("Ortalama Fiyat", f"{valid.mean():,.2f}"),
                ("Minimum Fiyat", f"{valid.min():,.2f}"),
                ("Maksimum Fiyat", f"{valid.max():,.2f}"),
            ])

    if "TEDARIKCI" in df.columns:
        summary.append(("", ""))
        summary.append(("Tedarikçi Bazlı Dağılım", ""))
        for name, count in df["TEDARIKCI"].value_counts().items():
            summary.append((str(name), count))

    label_font = Font(bold=True, size=11, color="1F4E79")
    label_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for row_idx, (label, value) in enumerate(summary, 1):
        cell_l = ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        if label:
            cell_l.font = label_font
            cell_l.fill = label_fill

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
